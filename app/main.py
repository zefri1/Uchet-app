from __future__ import annotations

from collections import defaultdict
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import List, Optional, Dict
import zipfile
import tempfile

from fastapi import Depends, FastAPI, Form, HTTPException, Request, UploadFile, File
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse, StreamingResponse
from openpyxl import Workbook
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import io
from sqlalchemy.exc import OperationalError
from sqlalchemy import func, select, delete, text
from sqlalchemy.orm import Session, joinedload
from .db import Base, SessionLocal, engine, get_db
from .models import (
    AllocationRule,
    BillingPeriod,
    ChargeAllocation,
    GeneratedDocument,
    LeasePlacement,
    PropertyObject,
    Tenant,
    UtilityCharge,
    TrashBin,
    Tariff,
    TenantPayment,
)
from .paths import BUNDLE_ROOT
from .services.calculations import close_period, get_charge_amount, quantize_money, recalculate_period, overlaps
from .services.documents import (
    generate_act_docx,
    generate_invoice_docx,
    generate_register_xlsx,
    generate_invoice_xlsx,
    generate_act_xlsx,
    _safe_name,
)



try:
    Base.metadata.create_all(bind=engine, checkfirst=True)
except OperationalError as exc:
    if "already exists" not in str(exc):
        raise

app = FastAPI(title="Utility Billing App")
RESOURCE_DIR = BUNDLE_ROOT / "app"
templates = Jinja2Templates(directory=str(RESOURCE_DIR / "templates"))
app.mount("/static", StaticFiles(directory=str(RESOURCE_DIR / "static")), name="static")

@app.middleware("http")
async def redirect_v2_middleware(request: Request, call_next):
    response = await call_next(request)
    if response.status_code == 303:
        redirect_to = request.query_params.get("redirect_to")
        referer = request.headers.get("referer", "")
        if redirect_to and redirect_to.startswith("/") and not redirect_to.startswith("//"):
            response.headers["Location"] = redirect_to
        elif "/v2/" in referer:
            old_location = response.headers.get("Location", "")
            if old_location == "/objects":
                response.headers["Location"] = "/v2/directory?tab=objects"
            elif old_location == "/tenants":
                response.headers["Location"] = "/v2/directory?tab=tenants"
            elif old_location == "/placements":
                response.headers["Location"] = "/v2/directory?tab=placements"
            elif "/billing" in old_location:
                import urllib.parse
                parsed = urllib.parse.urlparse(old_location)
                query = parsed.query
                response.headers["Location"] = f"/v2/billing?{query}" if query else "/v2/billing"
            elif "/documents" in old_location:
                import urllib.parse
                referer_parsed = urllib.parse.urlparse(referer)
                query = referer_parsed.query
                response.headers["Location"] = f"/v2/billing?{query}#step-5" if query else "/v2/billing#step-5"
            elif old_location == "/payments":
                response.headers["Location"] = "/v2/payments"
            elif old_location == "/charges":
                response.headers["Location"] = "/v2/billing"
            elif old_location == "/settings":
                response.headers["Location"] = "/v2/settings"
            elif old_location == "/trash":
                response.headers["Location"] = "/v2/settings?tab=trash"
    return response

@app.get("/favicon.ico", include_in_schema=False)
def favicon():
    ico_path = BUNDLE_ROOT / "app_icon.ico"
    if ico_path.exists():
        return FileResponse(ico_path)
    fallback_path = RESOURCE_DIR / "static" / "app_icon.png"
    if fallback_path.exists():
        return FileResponse(fallback_path)
    raise HTTPException(status_code=404)

MAX_OBJECTS = 10
MAX_TENANTS = 50
UTILITY_TYPES = ("heat", "electricity", "water", "cleaning")
TENANT_TYPES = ("ИП", "ООО")
UTILITY_LABELS = {
    "heat": "Тепло",
    "electricity": "Электричество",
    "water": "Вода",
    "cleaning": "Уборка",
}


@app.get("/__health")
def health_check():
    return {"app": "RealEstateUtilityApp"}


def to_decimal(value: Optional[str], default: Optional[Decimal] = None) -> Optional[Decimal]:
    if value in (None, ""):
        return default
    try:
        return Decimal(value)
    except InvalidOperation as exc:
        raise HTTPException(status_code=400, detail=f"Некорректное число: {value}") from exc


def require_choice(value: str, allowed: tuple[str, ...], field_name: str) -> None:
    if value not in allowed:
        raise HTTPException(status_code=400, detail=f"Недопустимое значение поля '{field_name}'.")


def require_positive(value: Decimal, field_name: str) -> None:
    if value < 0:
        raise HTTPException(status_code=400, detail=f"Поле '{field_name}' не может быть отрицательным.")


STATUS_LABELS = {
    "draft": "Черновик",
    "calculated": "Рассчитан",
    "closed": "Закрыт",
}

PERIOD_TYPE_LABELS = {
    "month": "Месяц",
    "range": "Диапазон",
}

ALLOCATION_MODE_LABELS = {
    "area": "По площади",
    "manual": "Вручную",
    "mixed": "Смешанное",
}

BASE_AREA_MODE_LABELS = {
    "active_leases": "Сумма активных площадей",
    "object_total": "Общая площадь объекта",
}

VALUE_TYPE_LABELS = {
    "percent": "Процент",
    "fixed": "Фиксированная сумма",
}


def render(request: Request, page: str, db: Session, status_code: int = 200, **context):
    summary = {
        "objects": db.scalar(select(func.count(PropertyObject.id))) or 0,
        "tenants": db.scalar(select(func.count(Tenant.id))) or 0,
        "placements": db.scalar(select(func.count(LeasePlacement.id))) or 0,
        "periods": db.scalar(select(func.count(BillingPeriod.id))) or 0,
    }
    base_context = {
        "request": request,
        "page": page,
        "summary": summary,
        "utility_labels": UTILITY_LABELS,
        "status_labels": STATUS_LABELS,
        "period_type_labels": PERIOD_TYPE_LABELS,
        "allocation_mode_labels": ALLOCATION_MODE_LABELS,
        "base_area_mode_labels": BASE_AREA_MODE_LABELS,
        "value_type_labels": VALUE_TYPE_LABELS,
        "today": date.today(),
    }
    base_context.update(context)
    return templates.TemplateResponse(request, page, base_context, status_code=status_code)


def render_billing_page(
    request: Request,
    db: Session,
    *,
    selected_period_id: Optional[int] = None,
    error: Optional[str] = None,
    status_code: int = 200,
):
    periods = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).all()
    if selected_period_id is not None:
        selected_period = db.get(BillingPeriod, selected_period_id)
    else:
        selected_id = request.query_params.get("period_id")
        selected_period = None
        if selected_id and selected_id.isdigit():
            selected_period = db.get(BillingPeriod, int(selected_id))
        if not selected_period and periods:
            selected_period = periods[0]
    
    allocations = []
    totals: dict[str, Decimal] = {}
    tenant_balances = {}
    tenant_objects = {}
    
    if selected_period:
        allocations = db.scalars(
            select(ChargeAllocation)
            .options(
                joinedload(ChargeAllocation.object),
                joinedload(ChargeAllocation.tenant),
                joinedload(ChargeAllocation.utility_charge),
            )
            .where(ChargeAllocation.billing_period_id == selected_period.id)
            .order_by(ChargeAllocation.object_id, ChargeAllocation.tenant_id)
        ).all()
        
        tenant_balances = get_tenant_balances(db, selected_period)
        tenant_objects = {t.id: t for t in db.scalars(select(Tenant)).all()}
        for tenant_id, bal in tenant_balances.items():
            if tenant_id in tenant_objects:
                t = tenant_objects[tenant_id]
                totals[t.display_name] = bal["allocated"]
                
    grouped_periods = get_grouped_documents(db)
    
    return render(
        request,
        "billing.html",
        db,
        status_code=status_code,
        periods=periods,
        selected_period=selected_period,
        allocations=allocations,
        totals={k: quantize_money(v) for k, v in totals.items()},
        tenant_balances=tenant_balances,
        tenant_objects=tenant_objects,
        grouped_periods=grouped_periods,
        error=error,
    )


def get_grouped_documents(db: Session):
    periods = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).all()
    documents = db.scalars(
        select(GeneratedDocument)
        .options(joinedload(GeneratedDocument.tenant), joinedload(GeneratedDocument.billing_period))
        .order_by(GeneratedDocument.created_at.desc())
    ).all()
    
    period_map = {}
    for period in periods:
        period_map[period.id] = {
            "period": period,
            "registers": [],
            "tenants": {}
        }
        
    for doc in documents:
        p_id = doc.billing_period_id
        if p_id not in period_map:
            continue
        
        if doc.document_type == "register" and doc.tenant_id is None:
            if not period_map[p_id]["registers"]:
                period_map[p_id]["registers"].append(doc)
        elif doc.tenant:
            t_id = doc.tenant_id
            t_map = period_map[p_id]["tenants"]
            if t_id not in t_map:
                t_map[t_id] = {
                    "tenant_name": doc.tenant.display_name,
                    "register": None,
                    "invoice_docx": None,
                    "invoice_xlsx": None,
                    "act_docx": None,
                    "act_xlsx": None
                }
            if doc.document_type == "register":
                if t_map[t_id]["register"] is None:
                    t_map[t_id]["register"] = doc
            elif doc.document_type == "invoice":
                if t_map[t_id]["invoice_docx"] is None:
                    t_map[t_id]["invoice_docx"] = doc
            elif doc.document_type == "invoice_xlsx":
                if t_map[t_id]["invoice_xlsx"] is None:
                    t_map[t_id]["invoice_xlsx"] = doc
            elif doc.document_type == "act":
                if t_map[t_id]["act_docx"] is None:
                    t_map[t_id]["act_docx"] = doc
            elif doc.document_type == "act_xlsx":
                if t_map[t_id]["act_xlsx"] is None:
                    t_map[t_id]["act_xlsx"] = doc
                
    result = []
    for p_id in sorted(period_map.keys(), key=lambda k: period_map[k]["period"].start_date, reverse=True):
        p_data = period_map[p_id]
        sorted_tenants = sorted(p_data["tenants"].values(), key=lambda x: x["tenant_name"])
        result.append({
            "period": p_data["period"],
            "registers": p_data["registers"],
            "tenants": sorted_tenants
        })
    return result


def render_documents_page(
    request: Request,
    db: Session,
    *,
    error: Optional[str] = None,
    form_state: Optional[dict[str, object]] = None,
    status_code: int = 200,
):
    periods = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).all()
    tenants = db.scalars(select(Tenant).order_by(Tenant.display_name)).all()
    grouped_periods = get_grouped_documents(db)
    return render(
        request,
        "documents.html",
        db,
        status_code=status_code,
        periods=periods,
        tenants=tenants,
        grouped_periods=grouped_periods,
        error=error,
        form_state=form_state or {},
    )


@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    if request.url.path == "/placements":
        form = await request.form()

        def _as_int(value):
            try:
                return int(value)
            except (TypeError, ValueError):
                return value

        form_state = {
            "object_id": _as_int(form.get("object_id")),
            "tenant_id": _as_int(form.get("tenant_id")),
            "rental_address": form.get("rental_address", ""),
            "occupied_area": form.get("occupied_area", ""),
            "start_date": form.get("start_date", ""),
            "end_date": form.get("end_date", ""),
            "is_active": form.get("is_active") in ("on", "true", "True", "1"),
            "rent_tariff": form.get("rent_tariff", ""),
            "status": form.get("status", "active"),
            "comment": form.get("comment", ""),
        }
        db = SessionLocal()
        try:
            return render_placements_page(
                request,
                db,
                error=str(exc.detail),
                form_state=form_state,
                status_code=exc.status_code,
            )
        finally:
            db.close()
    return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})


def recalculate_drafts(db: Session):
    draft_periods = db.scalars(select(BillingPeriod).where(BillingPeriod.status == "draft")).all()
    for p in draft_periods:
        try:
            recalculate_period(db, p.id)
        except Exception:
            try:
                db.execute(delete(ChargeAllocation).where(ChargeAllocation.billing_period_id == p.id))
                p.status = "draft"
                db.commit()
            except Exception:
                db.rollback()


def make_db_backup(db: Session):
    backup = {
        "objects": [{k: v for k, v in obj.__dict__.items() if k != "_sa_instance_state"} for obj in db.scalars(select(PropertyObject)).all()],
        "tenants": [{k: v for k, v in t.__dict__.items() if k != "_sa_instance_state"} for t in db.scalars(select(Tenant)).all()],
        "placements": [{k: v for k, v in p.__dict__.items() if k != "_sa_instance_state"} for p in db.scalars(select(LeasePlacement)).all()],
        "rules": [{k: v for k, v in r.__dict__.items() if k != "_sa_instance_state"} for r in db.scalars(select(AllocationRule)).all()],
        "charges": [{k: v for k, v in c.__dict__.items() if k != "_sa_instance_state"} for c in db.scalars(select(UtilityCharge)).all()],
        "allocations": [{k: v for k, v in a.__dict__.items() if k != "_sa_instance_state"} for a in db.scalars(select(ChargeAllocation)).all()],
        "periods": [{k: v for k, v in p.__dict__.items() if k != "_sa_instance_state"} for p in db.scalars(select(BillingPeriod)).all()]
    }
    from .paths import DATA_DIR
    backup_file = DATA_DIR / "import_backup.json"
    with open(backup_file, "w", encoding="utf-8") as f:
        json.dump(backup, f, cls=DbEncoder, ensure_ascii=False)


def restore_db_backup(db: Session) -> bool:
    from .paths import DATA_DIR
    backup_file = DATA_DIR / "import_backup.json"
    if not backup_file.exists():
        return False
        
    with open(backup_file, "r", encoding="utf-8") as f:
        backup = json.load(f)
        
    # Truncate all tables
    db.execute(delete(ChargeAllocation))
    db.execute(delete(UtilityCharge))
    db.execute(delete(AllocationRule))
    db.execute(delete(LeasePlacement))
    db.execute(delete(Tenant))
    db.execute(delete(PropertyObject))
    db.execute(delete(BillingPeriod))
    db.flush()
    
    # Restore
    for data in backup.get("objects", []):
        db.add(PropertyObject(
            id=data.get("id"),
            name=data.get("name"),
            address=data.get("address"),
            total_area=Decimal(data.get("total_area")),
            note=data.get("note")
        ))
    for data in backup.get("tenants", []):
        db.add(Tenant(
            id=data.get("id"),
            tenant_type=data.get("tenant_type"),
            display_name=data.get("display_name"),
            phone=data.get("phone"),
            note=data.get("note")
        ))
    db.flush()
    
    for data in backup.get("periods", []):
        db.add(BillingPeriod(
            id=data.get("id"),
            period_type=data.get("period_type"),
            month_label=data.get("month_label"),
            start_date=date.fromisoformat(data.get("start_date")),
            end_date=date.fromisoformat(data.get("end_date")),
            status=data.get("status")
        ))
    db.flush()
    
    for data in backup.get("placements", []):
        db.add(LeasePlacement(
            id=data.get("id"),
            object_id=data.get("object_id"),
            tenant_id=data.get("tenant_id"),
            rental_address=data.get("rental_address"),
            occupied_area=Decimal(data.get("occupied_area")),
            start_date=date.fromisoformat(data.get("start_date")),
            end_date=date.fromisoformat(data.get("end_date")) if data.get("end_date") else None,
            is_active=data.get("is_active")
        ))
    for data in backup.get("rules", []):
        db.add(AllocationRule(
            id=data.get("id"),
            object_id=data.get("object_id"),
            utility_type=data.get("utility_type"),
            mode=data.get("mode"),
            base_area_mode=data.get("base_area_mode"),
            tenant_id=data.get("tenant_id"),
            value_type=data.get("value_type"),
            value=Decimal(data.get("value")) if data.get("value") else None,
            is_active=data.get("is_active")
        ))
    for data in backup.get("charges", []):
        db.add(UtilityCharge(
            id=data.get("id"),
            object_id=data.get("object_id"),
            billing_period_id=data.get("billing_period_id"),
            utility_type=data.get("utility_type"),
            input_mode=data.get("input_mode"),
            allocation_mode=data.get("allocation_mode"),
            amount=Decimal(data.get("amount")) if data.get("amount") else None,
            meter_from=Decimal(data.get("meter_from")) if data.get("meter_from") else None,
            meter_to=Decimal(data.get("meter_to")) if data.get("meter_to") else None,
            tariff=Decimal(data.get("tariff")) if data.get("tariff") else None,
            unit_name=data.get("unit_name"),
            comment=data.get("comment")
        ))
    db.flush()
    
    for data in backup.get("allocations", []):
        db.add(ChargeAllocation(
            id=data.get("id"),
            billing_period_id=data.get("billing_period_id"),
            utility_charge_id=data.get("utility_charge_id"),
            object_id=data.get("object_id"),
            tenant_id=data.get("tenant_id"),
            base_area=Decimal(data.get("base_area")),
            share_value=Decimal(data.get("share_value")),
            amount=Decimal(data.get("amount")),
            mode=data.get("mode"),
            manual_override=data.get("manual_override"),
            audit_payload=data.get("audit_payload")
        ))
        
    db.commit()
    
    try:
        backup_file.unlink()
    except Exception:
        pass
    return True


def get_validation_warnings(db: Session) -> List[str]:
    warnings = []
    
    # 1. Sum of tenant areas exceeds total area for an object
    objects = db.scalars(select(PropertyObject)).all()
    for obj in objects:
        active_placements = db.scalars(
            select(LeasePlacement)
            .where(LeasePlacement.object_id == obj.id, LeasePlacement.is_active.is_(True))
        ).all()
        total_occupied = sum(p.occupied_area for p in active_placements)
        if total_occupied > obj.total_area:
            warnings.append(
                f"Сумма активных площадей арендаторов в объекте \"{obj.name}\" ({total_occupied} кв.м) превышает общую площадь здания ({obj.total_area} кв.м)."
            )
            
    # Get latest period to use for "current period" checks
    latest_period = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).first()
    if latest_period:
        # 2. Tenant has no placements at all
        tenants = db.scalars(select(Tenant)).all()
        for tenant in tenants:
            placement_count = db.scalar(
                select(func.count(LeasePlacement.id))
                .where(LeasePlacement.tenant_id == tenant.id)
            )
            if placement_count == 0:
                warnings.append(
                    f"У арендатора \"{tenant.display_name}\" не заведено ни одного договора размещения. Арендатору без помещения нельзя начислить аренду."
                )
                
        # 3. Object has a utility charge in the latest period but no rule configured
        charges = db.scalars(
            select(UtilityCharge)
            .options(joinedload(UtilityCharge.object))
            .where(UtilityCharge.billing_period_id == latest_period.id)
        ).all()
        warned = set()
        for charge in charges:
            key = (charge.object_id, charge.utility_type)
            if key in warned:
                continue
            has_rule = db.scalar(
                select(func.count(AllocationRule.id))
                .where(
                    AllocationRule.object_id == charge.object_id,
                    AllocationRule.utility_type == charge.utility_type,
                    AllocationRule.is_active.is_(True)
                )
            )
            if not has_rule:
                utility_name = UTILITY_LABELS.get(charge.utility_type, charge.utility_type)
                warnings.append(
                    f"Для объекта \"{charge.object.name}\" внесен счет за \"{utility_name}\" в текущем периоде, но не настроено ни одно правило распределения."
                )
                warned.add(key)
    return warnings


import json
from datetime import datetime, date
from decimal import Decimal

class DbEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        if isinstance(obj, Decimal):
            return str(obj)
        return super().default(obj)

def move_to_trash(db: Session, record, entity_type: str, display_name: str):
    data = {}
    for k, v in record.__dict__.items():
        if k == "_sa_instance_state":
            continue
        if hasattr(v, "_sa_instance_state") or (isinstance(v, list) and len(v) > 0 and hasattr(v[0], "_sa_instance_state")):
            continue
        data[k] = v
    trash_item = TrashBin(
        entity_type=entity_type,
        original_id=record.id,
        display_name=display_name,
        data_json=json.dumps(data, cls=DbEncoder)
    )
    db.add(trash_item)
    db.flush()


@app.get("/dashboard")
def dashboard_page(request: Request, db: Session = Depends(get_db)):
    objects_count = db.scalar(select(func.count(PropertyObject.id))) or 0
    tenants_count = db.scalar(select(func.count(Tenant.id))) or 0
    placements_count = db.scalar(select(func.count(LeasePlacement.id))) or 0
    
    latest_period = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).first()
    total_amount = Decimal("0.00")
    if latest_period:
        total_amount = db.scalar(
            select(func.sum(ChargeAllocation.amount))
            .where(ChargeAllocation.billing_period_id == latest_period.id)
        ) or Decimal("0.00")
        
    warnings = get_validation_warnings(db)
    from .paths import DATA_DIR
    has_backup = (DATA_DIR / "import_backup.json").exists()
        
    return render(
        request,
        "dashboard.html",
        db,
        objects_count=objects_count,
        tenants_count=tenants_count,
        placements_count=placements_count,
        latest_period=latest_period,
        total_amount=quantize_money(total_amount),
        warnings=warnings,
        has_backup=has_backup,
    )



@app.get("/objects")
def objects_page(request: Request, db: Session = Depends(get_db)):
    objects = db.scalars(select(PropertyObject).order_by(PropertyObject.name)).all()
    # Рассчитываем занятую площадь по каждому объекту
    occupied_areas: dict[int, Decimal] = {}
    for obj in objects:
        active_placements = db.scalars(
            select(LeasePlacement).where(
                LeasePlacement.object_id == obj.id,
                LeasePlacement.is_active.is_(True)
            )
        ).all()
        occupied_areas[obj.id] = sum(
            Decimal(p.occupied_area) for p in active_placements
        )
    return render(request, "objects.html", db, objects=objects, occupied_areas=occupied_areas)


@app.post("/objects")
def create_object(
    name: str = Form(...),
    address: str = Form(""),
    total_area: str = Form("0.00"),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    if (db.scalar(select(func.count(PropertyObject.id))) or 0) >= MAX_OBJECTS:
        raise HTTPException(status_code=400, detail="Нельзя добавить больше 10 объектов.")

    total_area_value = to_decimal(total_area, Decimal("0.00")) or Decimal("0.00")
    require_positive(total_area_value, "общая площадь")

    obj = PropertyObject(
        name=name.strip(),
        address=address.strip() if address and address.strip() else "Не указан",
        total_area=total_area_value,
        note=note.strip() or None,
    )
    db.add(obj)
    db.commit()
    return RedirectResponse("/objects", status_code=303)


@app.post("/objects/{object_id}/edit")
def edit_object(
    object_id: int,
    name: str = Form(...),
    address: str = Form(""),
    total_area: str = Form("0.00"),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    obj = db.get(PropertyObject, object_id)
    if obj is None:
        raise HTTPException(status_code=404, detail="Объект не найден.")
    total_area_value = to_decimal(total_area, Decimal("0.00")) or Decimal("0.00")
    require_positive(total_area_value, "общая площадь")
    obj.name = name.strip()
    obj.address = address.strip() if address and address.strip() else "Не указан"
    obj.total_area = total_area_value
    obj.note = note.strip() or None
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/objects", status_code=303)


@app.post("/objects/{object_id}/delete")
def delete_object(object_id: int, db: Session = Depends(get_db)):
    obj = db.get(PropertyObject, object_id)
    if obj is None:
        raise HTTPException(status_code=404, detail="Объект не найден.")
    move_to_trash(db, obj, "property_object", f"Объект: {obj.name} ({obj.total_area} кв.м)")
    db.execute(delete(ChargeAllocation).where(ChargeAllocation.object_id == object_id))
    db.execute(delete(UtilityCharge).where(UtilityCharge.object_id == object_id))
    db.execute(delete(LeasePlacement).where(LeasePlacement.object_id == object_id))
    db.execute(delete(AllocationRule).where(AllocationRule.object_id == object_id))
    db.execute(delete(Tariff).where(Tariff.object_id == object_id))
    db.delete(obj)
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/objects", status_code=303)


@app.get("/tenants")
def tenants_page(request: Request, db: Session = Depends(get_db)):
    tenants = db.scalars(select(Tenant).order_by(Tenant.display_name)).all()
    return render(request, "tenants.html", db, tenants=tenants)


@app.post("/tenants")
def create_tenant(
    tenant_type: str = Form(...),
    display_name: str = Form(...),
    phone: str = Form(""),
    initial_balance: str = Form("0.00"),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    if (db.scalar(select(func.count(Tenant.id))) or 0) >= MAX_TENANTS:
        raise HTTPException(status_code=400, detail="Нельзя добавить больше 50 арендаторов.")

    require_choice(tenant_type, TENANT_TYPES, "тип арендатора")
    balance_val = to_decimal(initial_balance, Decimal("0.00")) or Decimal("0.00")

    tenant = Tenant(
        tenant_type=tenant_type,
        display_name=display_name.strip(),
        phone=phone.strip() or None,
        initial_balance=balance_val,
        note=note.strip() or None,
    )
    db.add(tenant)
    db.commit()
    return RedirectResponse("/tenants", status_code=303)


@app.post("/tenants/{tenant_id}/edit")
def edit_tenant(
    tenant_id: int,
    tenant_type: str = Form(...),
    display_name: str = Form(...),
    phone: str = Form(""),
    initial_balance: str = Form("0.00"),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    tenant = db.get(Tenant, tenant_id)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Арендатор не найден.")
    require_choice(tenant_type, TENANT_TYPES, "тип арендатора")
    balance_val = to_decimal(initial_balance, Decimal("0.00")) or Decimal("0.00")
    tenant.tenant_type = tenant_type
    tenant.display_name = display_name.strip()
    tenant.phone = phone.strip() or None
    tenant.initial_balance = balance_val
    tenant.note = note.strip() or None
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/tenants", status_code=303)


@app.post("/tenants/{tenant_id}/delete")
def delete_tenant(tenant_id: int, db: Session = Depends(get_db)):
    tenant = db.get(Tenant, tenant_id)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Арендатор не найден.")
    move_to_trash(db, tenant, "tenant", f"Арендатор: {tenant.tenant_type} {tenant.display_name}")
    db.execute(delete(ChargeAllocation).where(ChargeAllocation.tenant_id == tenant_id))
    db.execute(delete(LeasePlacement).where(LeasePlacement.tenant_id == tenant_id))
    db.execute(delete(AllocationRule).where(AllocationRule.tenant_id == tenant_id))
    db.execute(delete(GeneratedDocument).where(GeneratedDocument.tenant_id == tenant_id))
    db.execute(delete(TenantPayment).where(TenantPayment.tenant_id == tenant_id))
    db.execute(delete(Tariff).where(Tariff.tenant_id == tenant_id))
    db.delete(tenant)
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/tenants", status_code=303)


@app.get("/placements")
def placements_page(request: Request, db: Session = Depends(get_db)):
    return render_placements_page(request, db)


def render_placements_page(
    request: Request,
    db: Session,
    *,
    error: Optional[str] = None,
    form_state: Optional[dict[str, object]] = None,
    status_code: int = 200,
):
    placements = db.scalars(
        select(LeasePlacement)
        .options(joinedload(LeasePlacement.object), joinedload(LeasePlacement.tenant))
        .order_by(LeasePlacement.start_date.desc())
    ).all()
    objects = db.scalars(select(PropertyObject).order_by(PropertyObject.name)).all()
    tenants = db.scalars(select(Tenant).order_by(Tenant.display_name)).all()
    return render(
        request,
        "placements.html",
        db,
        status_code=status_code,
        placements=placements,
        objects=objects,
        tenants=tenants,
        error=error,
        form_state=form_state or {},
    )


@app.post("/placements")
def create_placement(
    request: Request,
    object_id: int = Form(...),
    tenant_id: int = Form(...),
    rental_address: str = Form(...),
    occupied_area: str = Form(...),
    start_date: date = Form(...),
    end_date: Optional[date] = Form(None),
    is_active: bool = Form(False),
    rent_tariff: Optional[str] = Form(None),
    status: str = Form("active"),
    comment: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    obj = db.get(PropertyObject, object_id)
    if obj is None:
        raise HTTPException(status_code=404, detail="Объект не найден.")
    if db.get(Tenant, tenant_id) is None:
        raise HTTPException(status_code=404, detail="Арендатор не найден.")

    area = to_decimal(occupied_area, Decimal("0")) or Decimal("0")
    require_positive(area, "занимаемая площадь")
    
    rent_tariff_val = to_decimal(rent_tariff, None)
    if rent_tariff_val is not None:
        require_positive(rent_tariff_val, "тариф аренды")

    if area > Decimal(obj.total_area):
        return render_placements_page(
            request,
            db,
            error=f"Занимаемая площадь ({area}) не может быть больше общей площади объекта ({obj.total_area}).",
            form_state={
                "object_id": object_id,
                "tenant_id": tenant_id,
                "rental_address": rental_address,
                "occupied_area": occupied_area,
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat() if end_date else "",
                "is_active": is_active,
                "rent_tariff": rent_tariff,
                "status": status,
                "comment": comment,
            },
            status_code=400,
        )

    placement = LeasePlacement(
        object_id=object_id,
        tenant_id=tenant_id,
        rental_address=rental_address.strip(),
        occupied_area=area,
        start_date=start_date,
        end_date=end_date,
        is_active=is_active,
        rent_tariff=rent_tariff_val,
        status=status.strip(),
        comment=comment.strip() if comment else None,
    )
    db.add(placement)
    db.commit()
    # Recalculate draft periods
    draft_periods = db.scalars(select(BillingPeriod).where(BillingPeriod.status == "draft")).all()
    for p in draft_periods:
        try:
            recalculate_period(db, p.id)
        except Exception:
            pass
    return RedirectResponse("/placements", status_code=303)


@app.post("/placements/{placement_id}/edit")
def edit_placement(
    request: Request,
    placement_id: int,
    object_id: int = Form(...),
    tenant_id: int = Form(...),
    rental_address: str = Form(...),
    occupied_area: str = Form(...),
    start_date: date = Form(...),
    end_date: Optional[date] = Form(None),
    is_active: bool = Form(False),
    rent_tariff: Optional[str] = Form(None),
    status: str = Form("active"),
    comment: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    placement = db.get(LeasePlacement, placement_id)
    if placement is None:
        raise HTTPException(status_code=404, detail="Размещение не найдено.")
    obj = db.get(PropertyObject, object_id)
    if obj is None:
        raise HTTPException(status_code=404, detail="Объект не найден.")
    if db.get(Tenant, tenant_id) is None:
        raise HTTPException(status_code=404, detail="Арендатор не найден.")

    area = to_decimal(occupied_area, Decimal("0")) or Decimal("0")
    require_positive(area, "занимаемая площадь")
    
    rent_tariff_val = to_decimal(rent_tariff, None)
    if rent_tariff_val is not None:
        require_positive(rent_tariff_val, "тариф аренды")

    if area > Decimal(obj.total_area):
        return render_placements_page(
            request,
            db,
            error=f"Занимаемая площадь ({area}) не может быть больше общей площади объекта ({obj.total_area}).",
            status_code=400,
        )

    placement.object_id = object_id
    placement.tenant_id = tenant_id
    placement.rental_address = rental_address.strip()
    placement.occupied_area = area
    placement.start_date = start_date
    placement.end_date = end_date
    placement.is_active = is_active
    placement.rent_tariff = rent_tariff_val
    placement.status = status.strip()
    placement.comment = comment.strip() if comment else None
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/placements", status_code=303)


@app.post("/placements/{placement_id}/delete")
def delete_placement(placement_id: int, db: Session = Depends(get_db)):
    placement = db.get(LeasePlacement, placement_id)
    if placement is None:
        raise HTTPException(status_code=404, detail="Размещение не найдено.")
    move_to_trash(db, placement, "lease_placement", f"Размещение: {placement.tenant.display_name} -> {placement.object.name} ({placement.occupied_area} кв.м)")
    db.delete(placement)
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/placements", status_code=303)


@app.get("/charges")
def charges_page(request: Request, db: Session = Depends(get_db)):
    objects = db.scalars(select(PropertyObject).order_by(PropertyObject.name)).all()
    periods = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).all()
    tenants = db.scalars(select(Tenant).order_by(Tenant.display_name)).all()
    charges = db.scalars(
        select(UtilityCharge)
        .options(joinedload(UtilityCharge.object), joinedload(UtilityCharge.billing_period))
        .order_by(UtilityCharge.created_at.desc())
    ).all()
    rules = db.scalars(
        select(AllocationRule)
        .options(joinedload(AllocationRule.object), joinedload(AllocationRule.tenant))
        .order_by(AllocationRule.created_at.desc())
    ).all()
    tariffs = db.scalars(
        select(Tariff)
        .options(joinedload(Tariff.object), joinedload(Tariff.tenant))
        .order_by(Tariff.created_at.desc())
    ).all()
    return render(
        request,
        "charges.html",
        db,
        objects=objects,
        periods=periods,
        tenants=tenants,
        charges=charges,
        rules=rules,
        tariffs=tariffs,
    )


@app.post("/charges")
def create_charge(
    object_id: int = Form(...),
    billing_period_id: int = Form(...),
    utility_type: str = Form(...),
    input_mode: str = Form(...),
    allocation_mode: str = Form(...),
    amount: str = Form(""),
    meter_from: str = Form(""),
    meter_to: str = Form(""),
    tariff: str = Form(""),
    unit_name: str = Form(""),
    comment: str = Form(""),
    db: Session = Depends(get_db),
):
    if db.get(PropertyObject, object_id) is None:
        raise HTTPException(status_code=404, detail="Объект не найден.")
    if db.get(BillingPeriod, billing_period_id) is None:
        raise HTTPException(status_code=404, detail="Период не найден.")
    require_choice(utility_type, UTILITY_TYPES, "услуга")

    charge = UtilityCharge(
        object_id=object_id,
        billing_period_id=billing_period_id,
        utility_type=utility_type,
        input_mode=input_mode,
        allocation_mode=allocation_mode,
        amount=to_decimal(amount),
        meter_from=to_decimal(meter_from),
        meter_to=to_decimal(meter_to),
        tariff=to_decimal(tariff),
        unit_name=unit_name.strip() or None,
        comment=comment.strip() or None,
    )
    get_charge_amount(charge)
    db.add(charge)
    db.commit()
    try:
        recalculate_period(db, billing_period_id)
    except Exception:
        pass
    return RedirectResponse("/charges", status_code=303)


@app.post("/rules")
def create_rule(
    object_id: int = Form(...),
    utility_type: str = Form(...),
    mode: str = Form(...),
    base_area_mode: str = Form(...),
    tenant_id: str = Form(""),
    value_type: str = Form(""),
    value: str = Form(""),
    db: Session = Depends(get_db),
):
    if db.get(PropertyObject, object_id) is None:
        raise HTTPException(status_code=404, detail="Объект не найден.")
    require_choice(utility_type, UTILITY_TYPES, "услуга")
    
    actual_tenant_id: Optional[int] = None
    if tenant_id and tenant_id.strip():
        try:
            actual_tenant_id = int(tenant_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Недопустимый ID арендатора.")

    if actual_tenant_id and db.get(Tenant, actual_tenant_id) is None:
        raise HTTPException(status_code=404, detail="Арендатор не найден.")

    rule = AllocationRule(
        object_id=object_id,
        utility_type=utility_type,
        mode=mode,
        base_area_mode=base_area_mode,
        tenant_id=actual_tenant_id,
        value_type=value_type or None,
        value=to_decimal(value),
    )
    db.add(rule)
    db.commit()
    # Recalculate draft periods
    draft_periods = db.scalars(select(BillingPeriod).where(BillingPeriod.status == "draft")).all()
    for p in draft_periods:
        try:
            recalculate_period(db, p.id)
        except Exception:
            pass
    return RedirectResponse("/charges", status_code=303)


@app.post("/charges/{charge_id}/edit")
def edit_charge(
    charge_id: int,
    object_id: int = Form(...),
    billing_period_id: int = Form(...),
    utility_type: str = Form(...),
    input_mode: str = Form(...),
    allocation_mode: str = Form(...),
    amount: str = Form(""),
    meter_from: str = Form(""),
    meter_to: str = Form(""),
    tariff: str = Form(""),
    unit_name: str = Form(""),
    comment: str = Form(""),
    db: Session = Depends(get_db),
):
    charge = db.get(UtilityCharge, charge_id)
    if charge is None:
        raise HTTPException(status_code=404, detail="Счет не найден.")
    if db.get(PropertyObject, object_id) is None:
        raise HTTPException(status_code=404, detail="Объект не найден.")
    if db.get(BillingPeriod, billing_period_id) is None:
        raise HTTPException(status_code=404, detail="Период не найден.")
    require_choice(utility_type, UTILITY_TYPES, "услуга")
    
    charge.object_id = object_id
    charge.billing_period_id = billing_period_id
    charge.utility_type = utility_type
    charge.input_mode = input_mode
    charge.allocation_mode = allocation_mode
    charge.amount = to_decimal(amount)
    charge.meter_from = to_decimal(meter_from)
    charge.meter_to = to_decimal(meter_to)
    charge.tariff = to_decimal(tariff)
    charge.unit_name = unit_name.strip() or None
    charge.comment = comment.strip() or None
    
    get_charge_amount(charge)
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/charges", status_code=303)


@app.post("/charges/{charge_id}/delete")
def delete_charge(charge_id: int, db: Session = Depends(get_db)):
    charge = db.get(UtilityCharge, charge_id)
    if charge is None:
        raise HTTPException(status_code=404, detail="Счет не найден.")
    utility_name = UTILITY_LABELS.get(charge.utility_type, charge.utility_type)
    move_to_trash(db, charge, "utility_charge", f"Счет: {utility_name} на объекте {charge.object.name} ({charge.amount or 0} руб.)")
    db.execute(delete(ChargeAllocation).where(ChargeAllocation.utility_charge_id == charge_id))
    db.delete(charge)
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/charges", status_code=303)


@app.post("/rules/{rule_id}/edit")
def edit_rule(
    rule_id: int,
    object_id: int = Form(...),
    utility_type: str = Form(...),
    mode: str = Form(...),
    base_area_mode: str = Form(...),
    tenant_id: str = Form(""),
    value_type: str = Form(""),
    value: str = Form(""),
    db: Session = Depends(get_db),
):
    rule = db.get(AllocationRule, rule_id)
    if rule is None:
        raise HTTPException(status_code=404, detail="Правило не найдено.")
    if db.get(PropertyObject, object_id) is None:
        raise HTTPException(status_code=404, detail="Объект не найден.")
    require_choice(utility_type, UTILITY_TYPES, "услуга")
    
    actual_tenant_id: Optional[int] = None
    if tenant_id and tenant_id.strip():
        try:
            actual_tenant_id = int(tenant_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Недопустимый ID арендатора.")

    if actual_tenant_id and db.get(Tenant, actual_tenant_id) is None:
        raise HTTPException(status_code=404, detail="Арендатор не найден.")
        
    rule.object_id = object_id
    rule.utility_type = utility_type
    rule.mode = mode
    rule.base_area_mode = base_area_mode
    rule.tenant_id = actual_tenant_id
    rule.value_type = value_type or None
    rule.value = to_decimal(value)
    
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/charges", status_code=303)


@app.post("/rules/{rule_id}/delete")
def delete_rule(rule_id: int, db: Session = Depends(get_db)):
    rule = db.get(AllocationRule, rule_id)
    if rule is None:
        raise HTTPException(status_code=404, detail="Правило не найдено.")
    utility_name = UTILITY_LABELS.get(rule.utility_type, rule.utility_type)
    tenant_name = rule.tenant.display_name if rule.tenant else "всех"
    move_to_trash(db, rule, "allocation_rule", f"Правило: {utility_name} на {rule.object.name} для {tenant_name}")
    db.delete(rule)
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/charges", status_code=303)


@app.post("/tariffs")
def create_tariff(
    object_id: int = Form(...),
    tenant_id: str = Form(""),
    utility_type: str = Form(...),
    value: str = Form(...),
    unit_name: str = Form(""),
    start_date: date = Form(...),
    end_date: Optional[date] = Form(None),
    is_active: bool = Form(False),
    db: Session = Depends(get_db),
):
    if db.get(PropertyObject, object_id) is None:
        raise HTTPException(status_code=404, detail="Объект не найден.")
        
    actual_tenant_id: Optional[int] = None
    if tenant_id and tenant_id.strip():
        try:
            actual_tenant_id = int(tenant_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Недопустимый ID арендатора.")
        if db.get(Tenant, actual_tenant_id) is None:
            raise HTTPException(status_code=404, detail="Арендатор не найден.")

    allowed_types = ("rent",) + UTILITY_TYPES
    require_choice(utility_type, allowed_types, "тип тарифа")

    val = to_decimal(value, Decimal("0")) or Decimal("0")
    require_positive(val, "значение тарифа")

    tariff = Tariff(
        object_id=object_id,
        tenant_id=actual_tenant_id,
        utility_type=utility_type,
        value=val,
        unit_name=unit_name.strip() or None,
        start_date=start_date,
        end_date=end_date,
        is_active=is_active,
    )
    db.add(tariff)
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/charges", status_code=303)


@app.post("/tariffs/{tariff_id}/edit")
def edit_tariff(
    tariff_id: int,
    object_id: int = Form(...),
    tenant_id: str = Form(""),
    utility_type: str = Form(...),
    value: str = Form(...),
    unit_name: str = Form(""),
    start_date: date = Form(...),
    end_date: Optional[date] = Form(None),
    is_active: bool = Form(False),
    db: Session = Depends(get_db),
):
    tariff = db.get(Tariff, tariff_id)
    if tariff is None:
        raise HTTPException(status_code=404, detail="Тариф не найден.")
    if db.get(PropertyObject, object_id) is None:
        raise HTTPException(status_code=404, detail="Объект не найден.")
        
    actual_tenant_id: Optional[int] = None
    if tenant_id and tenant_id.strip():
        try:
            actual_tenant_id = int(tenant_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Недопустимый ID арендатора.")
        if db.get(Tenant, actual_tenant_id) is None:
            raise HTTPException(status_code=404, detail="Арендатор не найден.")

    allowed_types = ("rent",) + UTILITY_TYPES
    require_choice(utility_type, allowed_types, "тип тарифа")

    val = to_decimal(value, Decimal("0")) or Decimal("0")
    require_positive(val, "значение тарифа")

    tariff.object_id = object_id
    tariff.tenant_id = actual_tenant_id
    tariff.utility_type = utility_type
    tariff.value = val
    tariff.unit_name = unit_name.strip() or None
    tariff.start_date = start_date
    tariff.end_date = end_date
    tariff.is_active = is_active
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/charges", status_code=303)


@app.post("/tariffs/{tariff_id}/delete")
def delete_tariff(tariff_id: int, db: Session = Depends(get_db)):
    tariff = db.get(Tariff, tariff_id)
    if tariff is None:
        raise HTTPException(status_code=404, detail="Тариф не найден.")
    utility_name = "Аренда" if tariff.utility_type == "rent" else UTILITY_LABELS.get(tariff.utility_type, tariff.utility_type)
    tenant_name = tariff.tenant.display_name if tariff.tenant else "всех"
    move_to_trash(db, tariff, "tariff", f"Тариф: {utility_name} на {tariff.object.name} для {tenant_name} ({tariff.value})")
    db.delete(tariff)
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/charges", status_code=303)


def get_tenant_balances(db: Session, period: BillingPeriod) -> dict[int, dict]:
    prev_periods = db.scalars(
        select(BillingPeriod).where(BillingPeriod.start_date < period.start_date)
    ).all()
    prev_period_ids = [p.id for p in prev_periods]

    tenants = db.scalars(select(Tenant)).all()
    balances = {}

    for t in tenants:
        initial = Decimal(t.initial_balance)
        prev_allocs = Decimal("0.00")
        if prev_period_ids:
            prev_allocs = db.scalar(
                select(func.sum(ChargeAllocation.amount))
                .where(ChargeAllocation.tenant_id == t.id)
                .where(ChargeAllocation.billing_period_id.in_(prev_period_ids))
            ) or Decimal("0.00")

        prev_pays = Decimal("0.00")
        if prev_period_ids:
            prev_pays = db.scalar(
                select(func.sum(TenantPayment.amount))
                .where(TenantPayment.tenant_id == t.id)
                .where(TenantPayment.is_active == True)
                .where(TenantPayment.billing_period_id.in_(prev_period_ids))
            ) or Decimal("0.00")

        incoming = initial + prev_allocs - prev_pays

        curr_allocs = db.scalar(
            select(func.sum(ChargeAllocation.amount))
            .where(ChargeAllocation.tenant_id == t.id)
            .where(ChargeAllocation.billing_period_id == period.id)
        ) or Decimal("0.00")

        curr_pays = db.scalar(
            select(func.sum(TenantPayment.amount))
            .where(TenantPayment.tenant_id == t.id)
            .where(TenantPayment.is_active == True)
            .where(TenantPayment.billing_period_id == period.id)
        ) or Decimal("0.00")

        outgoing = incoming + curr_allocs - curr_pays

        balances[t.id] = {
            "incoming": incoming,
            "allocated": curr_allocs,
            "paid": curr_pays,
            "outgoing": outgoing
        }
    return balances

@app.get("/billing")
def billing_page(request: Request, db: Session = Depends(get_db)):
    periods = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).all()
    selected_id = request.query_params.get("period_id")
    selected_period = db.get(BillingPeriod, int(selected_id)) if selected_id else (periods[0] if periods else None)
    allocations = []
    totals: dict[str, Decimal] = {}
    tenant_balances = {}
    tenant_objects = {}
    if selected_period:
        allocations = db.scalars(
            select(ChargeAllocation)
            .options(
                joinedload(ChargeAllocation.object),
                joinedload(ChargeAllocation.tenant),
                joinedload(ChargeAllocation.utility_charge),
            )
            .where(ChargeAllocation.billing_period_id == selected_period.id)
            .order_by(ChargeAllocation.object_id, ChargeAllocation.tenant_id)
        ).all()
        tenant_balances = get_tenant_balances(db, selected_period)
        tenant_objects = {t.id: t for t in db.scalars(select(Tenant)).all()}
        for tenant_id, bal in tenant_balances.items():
            t = tenant_objects[tenant_id]
            totals[t.display_name] = bal["allocated"]
            
    return render(
        request,
        "billing.html",
        db,
        periods=periods,
        selected_period=selected_period,
        allocations=allocations,
        totals={k: quantize_money(v) for k, v in totals.items()},
        tenant_balances=tenant_balances,
        tenant_objects=tenant_objects,
    )


@app.post("/billing/periods")
def create_billing_period(
    period_type: str = Form(...),
    month_label: str = Form(""),
    start_date: date = Form(...),
    end_date: date = Form(...),
    db: Session = Depends(get_db),
):
    if end_date < start_date:
        raise HTTPException(status_code=400, detail="Дата окончания не может быть раньше даты начала.")
    period = BillingPeriod(
        period_type=period_type,
        month_label=month_label.strip() or None,
        start_date=start_date,
        end_date=end_date,
    )
    db.add(period)
    db.commit()
    db.refresh(period)
    return RedirectResponse(f"/billing?period_id={period.id}", status_code=303)


@app.post("/billing/{period_id}/calculate")
def calculate_period(request: Request, period_id: int, db: Session = Depends(get_db)):
    warning = None
    try:
        recalculate_period(db, period_id)
    except ValueError as exc:
        # Расчёт завершён, но есть предупреждения (например, нет тарифа или нет арендаторов)
        # Отображаем их как warning, не блокируя результат
        warning = str(exc)
    redirect_to = request.query_params.get("redirect_to")
    if redirect_to:
        import urllib.parse
        url = f"{redirect_to}?period_id={period_id}"
        if warning:
            url += f"&error={urllib.parse.quote(warning)}"
        elif request.query_params.get("error"):
            url += f"&error={urllib.parse.quote(request.query_params.get('error'))}"
        return RedirectResponse(url, status_code=303)
        
    return render_billing_page(request, db, selected_period_id=period_id, error=warning,
                                status_code=200 if warning else 200)


@app.post("/billing/allocations/{allocation_id}")
def update_allocation(
    allocation_id: int,
    amount: str = Form(...),
    db: Session = Depends(get_db),
):
    allocation = db.get(ChargeAllocation, allocation_id)
    if allocation is None:
        raise HTTPException(status_code=404, detail="Начисление не найдено.")
    period = db.get(BillingPeriod, allocation.billing_period_id)
    if period and period.status == "closed":
        raise HTTPException(status_code=400, detail="Закрытый период нельзя редактировать.")
    allocation.amount = to_decimal(amount, Decimal("0")) or Decimal("0")
    allocation.manual_override = True
    db.commit()
    return RedirectResponse(f"/billing?period_id={allocation.billing_period_id}", status_code=303)


@app.post("/billing/{period_id}/close")
def close_billing_period(period_id: int, db: Session = Depends(get_db)):
    close_period(db, period_id)
    return RedirectResponse(f"/billing?period_id={period_id}", status_code=303)


@app.post("/billing/{period_id}/reopen")
def reopen_billing_period(period_id: int, db: Session = Depends(get_db)):
    period = db.get(BillingPeriod, period_id)
    if period is None:
        raise HTTPException(status_code=404, detail="Период не найден.")
    period.status = "draft"
    period.closed_at = None
    db.commit()
    return RedirectResponse(f"/billing?period_id={period_id}", status_code=303)


@app.post("/billing/{period_id}/reset")
def reset_billing_period(period_id: int, db: Session = Depends(get_db)):
    period = db.get(BillingPeriod, period_id)
    if period is None:
        raise HTTPException(status_code=404, detail="Период не найден.")
    if period.status == "closed":
        raise HTTPException(status_code=400, detail="Закрытый период нельзя сбросить.")
    
    db.execute(delete(ChargeAllocation).where(ChargeAllocation.billing_period_id == period_id))
    period.status = "draft"
    db.commit()
    
    try:
        recalculate_period(db, period_id)
    except ValueError:
        pass
        
    return RedirectResponse(f"/billing?period_id={period_id}", status_code=303)


@app.get("/documents")
def documents_page(request: Request, db: Session = Depends(get_db)):
    periods = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).all()
    tenants = db.scalars(select(Tenant).order_by(Tenant.display_name)).all()
    grouped_periods = get_grouped_documents(db)
    return render(request, "documents.html", db, periods=periods, tenants=tenants, grouped_periods=grouped_periods)


@app.get("/guide")
def guide_page(request: Request, db: Session = Depends(get_db)):
    return render(request, "guide.html", db)


@app.get("/trash")
def trash_page(request: Request, db: Session = Depends(get_db)):
    items = db.scalars(select(TrashBin).order_by(TrashBin.deleted_at.desc())).all()
    formatted_items = []
    ENTITY_TYPE_LABELS = {
        "property_object": "Объект недвижимости",
        "tenant": "Арендатор",
        "lease_placement": "Размещение арендатора",
        "utility_charge": "Коммунальный счет",
        "allocation_rule": "Правило распределения",
        "tariff": "Тариф",
        "tenant_payment": "Оплата арендатора",
    }
    for item in items:
        formatted_items.append({
            "id": item.id,
            "display_name": item.display_name,
            "entity_type_label": ENTITY_TYPE_LABELS.get(item.entity_type, item.entity_type),
            "deleted_at": item.deleted_at.strftime("%d.%m.%Y %H:%M")
        })
    return render(request, "trash.html", db, items=formatted_items)


@app.post("/trash/{trash_id}/restore")
def restore_item(trash_id: int, db: Session = Depends(get_db)):
    item = db.get(TrashBin, trash_id)
    if item is None:
        raise HTTPException(status_code=404, detail="Запись в корзине не найдена.")
    
    data = json.loads(item.data_json)
    
    try:
        if item.entity_type == "property_object":
            obj = PropertyObject(
                id=data.get("id"),
                name=data.get("name"),
                address=data.get("address"),
                total_area=Decimal(data.get("total_area")),
                note=data.get("note")
            )
            db.add(obj)
        elif item.entity_type == "tenant":
            tenant = Tenant(
                id=data.get("id"),
                tenant_type=data.get("tenant_type"),
                display_name=data.get("display_name"),
                phone=data.get("phone"),
                initial_balance=Decimal(data.get("initial_balance")) if data.get("initial_balance") is not None else Decimal("0.00"),
                note=data.get("note")
            )
            db.add(tenant)
        elif item.entity_type == "tenant_payment":
            tenant_exists = db.get(Tenant, data.get("tenant_id"))
            period_exists = db.get(BillingPeriod, data.get("billing_period_id"))
            if not tenant_exists or not period_exists:
                raise HTTPException(status_code=400, detail="Невозможно восстановить: связанный арендатор или период был удален.")
            
            pay = TenantPayment(
                id=data.get("id"),
                tenant_id=data.get("tenant_id"),
                billing_period_id=data.get("billing_period_id"),
                amount=Decimal(data.get("amount")),
                payment_date=date.fromisoformat(data.get("payment_date")),
                comment=data.get("comment"),
                is_active=data.get("is_active")
            )
            db.add(pay)
        elif item.entity_type == "lease_placement":
            obj_exists = db.get(PropertyObject, data.get("object_id"))
            tenant_exists = db.get(Tenant, data.get("tenant_id"))
            if not obj_exists or not tenant_exists:
                raise HTTPException(status_code=400, detail="Невозможно восстановить: связанный объект или арендатор был удален насовсем.")
            
            placement = LeasePlacement(
                id=data.get("id"),
                object_id=data.get("object_id"),
                tenant_id=data.get("tenant_id"),
                rental_address=data.get("rental_address"),
                occupied_area=Decimal(data.get("occupied_area")),
                start_date=date.fromisoformat(data.get("start_date")),
                end_date=date.fromisoformat(data.get("end_date")) if data.get("end_date") else None,
                is_active=data.get("is_active")
            )
            db.add(placement)
        elif item.entity_type == "utility_charge":
            obj_exists = db.get(PropertyObject, data.get("object_id"))
            period_exists = db.get(BillingPeriod, data.get("billing_period_id"))
            if not obj_exists or not period_exists:
                raise HTTPException(status_code=400, detail="Невозможно восстановить: связанный объект или период был удален.")
                
            charge = UtilityCharge(
                id=data.get("id"),
                object_id=data.get("object_id"),
                billing_period_id=data.get("billing_period_id"),
                utility_type=data.get("utility_type"),
                input_mode=data.get("input_mode"),
                allocation_mode=data.get("allocation_mode"),
                amount=Decimal(data.get("amount")) if data.get("amount") else None,
                meter_from=Decimal(data.get("meter_from")) if data.get("meter_from") else None,
                meter_to=Decimal(data.get("meter_to")) if data.get("meter_to") else None,
                tariff=Decimal(data.get("tariff")) if data.get("tariff") else None,
                unit_name=data.get("unit_name"),
                comment=data.get("comment")
            )
            db.add(charge)
        elif item.entity_type == "allocation_rule":
            obj_exists = db.get(PropertyObject, data.get("object_id"))
            if not obj_exists:
                raise HTTPException(status_code=400, detail="Невозможно восстановить: связанный объект был удален.")
            
            rule = AllocationRule(
                id=data.get("id"),
                object_id=data.get("object_id"),
                utility_type=data.get("utility_type"),
                mode=data.get("mode"),
                base_area_mode=data.get("base_area_mode"),
                tenant_id=data.get("tenant_id"),
                value_type=data.get("value_type"),
                value=Decimal(data.get("value")) if data.get("value") else None,
                is_active=data.get("is_active")
            )
            db.add(rule)
        elif item.entity_type == "tariff":
            obj_exists = db.get(PropertyObject, data.get("object_id"))
            if not obj_exists:
                raise HTTPException(status_code=400, detail="Невозможно восстановить: связанный объект был удален.")
            
            tariff = Tariff(
                id=data.get("id"),
                object_id=data.get("object_id"),
                tenant_id=data.get("tenant_id"),
                utility_type=data.get("utility_type"),
                value=Decimal(data.get("value")),
                unit_name=data.get("unit_name"),
                start_date=date.fromisoformat(data.get("start_date")),
                end_date=date.fromisoformat(data.get("end_date")) if data.get("end_date") else None,
                is_active=data.get("is_active")
            )
            db.add(tariff)
            
        db.delete(item)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Ошибка восстановления: {str(e)}")
        
    recalculate_drafts(db)
    return RedirectResponse("/trash", status_code=303)


@app.post("/trash/{trash_id}/delete")
def permanent_delete_item(trash_id: int, db: Session = Depends(get_db)):
    item = db.get(TrashBin, trash_id)
    if item is None:
        raise HTTPException(status_code=404, detail="Запись в корзине не найдена.")
    db.delete(item)
    db.commit()
    return RedirectResponse("/trash", status_code=303)


@app.post("/trash/clear")
def clear_trash(db: Session = Depends(get_db)):
    db.execute(delete(TrashBin))
    db.commit()
    return RedirectResponse("/trash", status_code=303)



@app.post("/documents/generate")
def generate_documents(
    request: Request,
    billing_period_id: int = Form(...),
    tenant_id: str = Form(""),
    scope: str = Form(...),
    db: Session = Depends(get_db),
):
    period = db.get(BillingPeriod, billing_period_id)
    if period is None:
        raise HTTPException(status_code=404, detail="Период не найден.")
        
    actual_tenant_id: Optional[int] = None
    if tenant_id and tenant_id.strip():
        try:
            actual_tenant_id = int(tenant_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Недопустимый ID арендатора.")

    try:
        if scope == "all":
            generate_register_xlsx(db, period)
            tenant_ids = db.scalars(
                select(ChargeAllocation.tenant_id).where(ChargeAllocation.billing_period_id == period.id).distinct()
            ).all()
            for current_tenant_id in tenant_ids:
                tenant = db.get(Tenant, current_tenant_id)
                if tenant:
                    generate_invoice_docx(db, period, tenant)
                    generate_act_docx(db, period, tenant)
                    generate_invoice_xlsx(db, period, tenant)
                    generate_act_xlsx(db, period, tenant)
        else:
            tenant = db.get(Tenant, actual_tenant_id) if actual_tenant_id else None
            if tenant is None:
                raise ValueError("Для режима 'Для одного арендатора' необходимо выбрать конкретного арендатора.")
            generate_invoice_docx(db, period, tenant)
            generate_act_docx(db, period, tenant)
            generate_invoice_xlsx(db, period, tenant)
            generate_act_xlsx(db, period, tenant)
            generate_register_xlsx(db, period, tenant)
    except ValueError as exc:
        return render_documents_page(
            request,
            db,
            error=str(exc),
            form_state={
                "billing_period_id": billing_period_id,
                "tenant_id": tenant_id,
                "scope": scope,
            },
            status_code=400,
        )
    return RedirectResponse("/documents", status_code=303)


@app.get("/api/periods/{period_id}/allocations")
def api_period_allocations(period_id: int, db: Session = Depends(get_db)):
    """Возвращает JSON со списком арендаторов и их начислениями за период."""
    period = db.get(BillingPeriod, period_id)
    if not period:
        raise HTTPException(status_code=404, detail="Период не найден")

    UTILITY_LABELS = {
        "heat": "Тепло",
        "electricity": "Электричество",
        "water": "Водоснабжение",
        "cleaning": "Уборка",
        "other": "Прочее",
    }

    allocs = db.scalars(
        select(ChargeAllocation)
        .options(
            joinedload(ChargeAllocation.tenant),
            joinedload(ChargeAllocation.utility_charge),
        )
        .where(ChargeAllocation.billing_period_id == period_id)
        .order_by(ChargeAllocation.tenant_id)
    ).all()

    by_tenant: dict = defaultdict(list)
    for a in allocs:
        by_tenant[a.tenant_id].append(a)

    result = []
    for tenant_id, items in by_tenant.items():
        tenant = items[0].tenant
        services: dict = defaultdict(float)
        for a in items:
            # utility_type живёт в связанном UtilityCharge, не в ChargeAllocation
            utype = a.utility_charge.utility_type if a.utility_charge else "other"
            label = UTILITY_LABELS.get(utype, utype)
            services[label] += float(a.amount or 0)
        result.append({
            "tenant_id": tenant_id,
            "tenant_name": tenant.display_name if tenant else f"ID {tenant_id}",
            "total": round(sum(services.values()), 2),
            "services": dict(services),
        })

    return JSONResponse(result)


@app.post("/documents/generate-selected")
def generate_documents_selected(
    request: Request,
    billing_period_id: int = Form(...),
    tenant_ids: List[str] = Form(default=[]),
    db: Session = Depends(get_db),
):
    """Генерирует документы для выбранных арендаторов из списка начислений."""
    period = db.get(BillingPeriod, billing_period_id)
    if period is None:
        raise HTTPException(status_code=404, detail="Период не найден.")

    if not tenant_ids:
        return RedirectResponse("/documents", status_code=303)

    try:
        for tid_str in tenant_ids:
            try:
                tenant = db.get(Tenant, int(tid_str))
            except ValueError:
                continue
            if tenant:
                generate_invoice_docx(db, period, tenant)
                generate_act_docx(db, period, tenant)
                generate_invoice_xlsx(db, period, tenant)
                generate_act_xlsx(db, period, tenant)
                generate_register_xlsx(db, period, tenant)

        generate_register_xlsx(db, period)
    except ValueError as exc:
        return render_documents_page(
            request,
            db,
            error=str(exc),
            form_state={
                "billing_period_id": billing_period_id,
            },
            status_code=400,
        )
    return RedirectResponse("/documents", status_code=303)


def _get_clean_document_filename(doc: GeneratedDocument) -> str:
    p = doc.billing_period
    period_str = p.month_label or f"{p.start_date.isoformat()}_{p.end_date.isoformat()}"
    tenant_str = ""
    if doc.tenant:
        tenant_str = f"_{_safe_name(doc.tenant.display_name)}"
    
    ext = "docx"
    if doc.document_type.endswith("_xlsx") or doc.document_type == "register":
        ext = "xlsx"
        
    if doc.document_type == "register":
        if doc.tenant_id:
            return f"Реестр{tenant_str}_{period_str}.xlsx"
        return f"Реестр_начислений_{period_str}.xlsx"
        
    doc_type_label = "Счет" if "invoice" in doc.document_type else "Акт"
    return f"{doc_type_label}{tenant_str}_{period_str}.{ext}"


@app.get("/documents/{document_id}/download")
def download_document(document_id: int, db: Session = Depends(get_db)):
    document = db.scalar(
        select(GeneratedDocument)
        .options(joinedload(GeneratedDocument.tenant), joinedload(GeneratedDocument.billing_period))
        .where(GeneratedDocument.id == document_id)
    )
    if document is None:
        raise HTTPException(status_code=404, detail="Документ не найден.")
    clean_name = _get_clean_document_filename(document)
    import os
    if not os.path.exists(document.file_path):
        import urllib.parse
        error_msg = urllib.parse.quote("Файл документа отсутствует на диске. Пожалуйста, сформируйте документы заново.")
        return RedirectResponse(f"/v2/billing?period_id={document.billing_period_id}&error={error_msg}", status_code=303)
        
    import urllib.parse
    encoded_name = urllib.parse.quote(clean_name)
    headers = {"Content-Disposition": f"attachment; filename*=utf-8''{encoded_name}"}
    return FileResponse(document.file_path, headers=headers)


@app.get("/documents/periods/{period_id}/zip")
def download_period_zip(period_id: int, db: Session = Depends(get_db)):
    period = db.get(BillingPeriod, period_id)
    if period is None:
        raise HTTPException(status_code=404, detail="Период не найден.")
        
    # Get all documents ordered by created_at desc so the newest are processed first
    documents = db.scalars(
        select(GeneratedDocument)
        .options(joinedload(GeneratedDocument.tenant), joinedload(GeneratedDocument.billing_period))
        .where(GeneratedDocument.billing_period_id == period.id)
        .order_by(GeneratedDocument.created_at.desc())
    ).all()
    
    if not documents:
        raise HTTPException(status_code=400, detail="Нет документов для скачивания за этот период.")
        
    # Deduplicate: only keep the newest document for each combination of (document_type, tenant_id)
    latest_docs = {}
    for doc in documents:
        key = (doc.document_type, doc.tenant_id)
        if key not in latest_docs:
            latest_docs[key] = doc
            
    final_docs = list(latest_docs.values())
        
    temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
    temp_zip.close()
    
    with zipfile.ZipFile(temp_zip.name, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for doc in final_docs:
            file_path = Path(doc.file_path)
            if file_path.exists():
                clean_name = _get_clean_document_filename(doc)
                zip_file.write(file_path, arcname=clean_name)
                
    period_label = period.month_label or f"{period.start_date}_{period.end_date}"
    zip_name = f"documents_{period_label}.zip"
    
    import urllib.parse
    encoded_zip_name = urllib.parse.quote(zip_name)
    headers = {"Content-Disposition": f"attachment; filename*=utf-8''{encoded_zip_name}"}
    return FileResponse(
        temp_zip.name,
        media_type="application/zip",
        headers=headers
    )


@app.get("/import/template")
def download_import_template():
    from openpyxl import Workbook
    workbook = Workbook()
    
    # Sheet 1: Objects
    sheet_obj = workbook.active
    sheet_obj.title = "Объекты"
    sheet_obj.append(["Название", "Адрес", "Площадь, кв.м", "Примечание"])
    sheet_obj.append(["БЦ Север", "Москва, ул. Ленина, д. 5", 500.50, "Главный корпус"])
    sheet_obj.append(["Склад Юг", "Москва, ул. Южная, д. 12", 250.00, "Теплый склад"])
    
    # Sheet 2: Tenants
    sheet_tenant = workbook.create_sheet(title="Арендаторы")
    sheet_tenant.append(["Тип (ИП/ООО)", "ФИО / Наименование", "Телефон", "Примечание"])
    sheet_tenant.append(["ООО", "Ромашка", "+7 (999) 111-22-33", "Офис 101"])
    sheet_tenant.append(["ИП", "Иванов Иван Иванович", "+7 (999) 555-44-33", "Офис 102"])
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    temp_file.close()
    workbook.save(temp_file.name)
    
    return FileResponse(
        temp_file.name,
        filename="import_template.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.post("/import")
def import_excel_data(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    # 0. Create database backup in JSON before running import
    make_db_backup(db)
    
    from openpyxl import load_workbook
    try:
        contents = file.file.read()
        buffer = io.BytesIO(contents)
        workbook = load_workbook(buffer, data_only=True)
        
        objects_created = 0
        tenants_created = 0
        
        # 1. Parse objects
        if "Объекты" in workbook.sheetnames:
            sheet = workbook["Объекты"]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                name, address, area, note = row[0], row[1] or "", row[2] or 0, row[3]
                
                existing_count = db.scalar(select(func.count(PropertyObject.id))) or 0
                if existing_count >= MAX_OBJECTS:
                    continue
                    
                obj = PropertyObject(
                    name=str(name).strip(),
                    address=str(address).strip(),
                    total_area=Decimal(str(area)),
                    note=str(note).strip() if note else None
                )
                db.add(obj)
                objects_created += 1
                
        # 2. Parse tenants
        if "Арендаторы" in workbook.sheetnames:
            sheet = workbook["Арендаторы"]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0] or not row[1]:
                    continue
                t_type, display_name, phone, note = row[0], row[1], row[2], row[3]
                
                existing_count = db.scalar(select(func.count(Tenant.id))) or 0
                if existing_count >= MAX_TENANTS:
                    continue
                    
                tenant = Tenant(
                    tenant_type=str(t_type).strip(),
                    display_name=str(display_name).strip(),
                    phone=str(phone).strip() if phone else None,
                    note=str(note).strip() if note else None
                )
                db.add(tenant)
                tenants_created += 1
                
        db.commit()
        
        # Recalculate draft periods
        draft_periods = db.scalars(select(BillingPeriod).where(BillingPeriod.status == "draft")).all()
        for p in draft_periods:
            try:
                recalculate_period(db, p.id)
            except Exception:
                pass
                
        objects_count = db.scalar(select(func.count(PropertyObject.id))) or 0
        tenants_count = db.scalar(select(func.count(Tenant.id))) or 0
        placements_count = db.scalar(select(func.count(LeasePlacement.id))) or 0
        
        latest_period = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).first()
        total_amount = Decimal("0.00")
        if latest_period:
            total_amount = db.scalar(
                select(func.sum(ChargeAllocation.amount))
                .where(ChargeAllocation.billing_period_id == latest_period.id)
            ) or Decimal("0.00")
            
        warnings = get_validation_warnings(db)
        from .paths import DATA_DIR
        has_backup = (DATA_DIR / "import_backup.json").exists()
        success_msg = f"Импорт завершен успешно! Создано объектов: {objects_created}, арендаторов: {tenants_created}."
        return render(
            request,
            "dashboard.html",
            db,
            objects_count=objects_count,
            tenants_count=tenants_count,
            placements_count=placements_count,
            latest_period=latest_period,
            total_amount=quantize_money(total_amount),
            success_msg=success_msg,
            warnings=warnings,
            has_backup=has_backup,
        )
    except Exception as e:
        objects_count = db.scalar(select(func.count(PropertyObject.id))) or 0
        tenants_count = db.scalar(select(func.count(Tenant.id))) or 0
        placements_count = db.scalar(select(func.count(LeasePlacement.id))) or 0
        latest_period = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).first()
        total_amount = Decimal("0.00")
        if latest_period:
            total_amount = db.scalar(
                select(func.sum(ChargeAllocation.amount))
                .where(ChargeAllocation.billing_period_id == latest_period.id)
            ) or Decimal("0.00")
            
        warnings = get_validation_warnings(db)
        from .paths import DATA_DIR
        has_backup = (DATA_DIR / "import_backup.json").exists()
        return render(
            request,
            "dashboard.html",
            db,
            objects_count=objects_count,
            tenants_count=tenants_count,
            placements_count=placements_count,
            latest_period=latest_period,
            total_amount=quantize_money(total_amount),
            error=f"Ошибка при импорте: {str(e)}",
            warnings=warnings,
            has_backup=has_backup,
            status_code=400,
        )


@app.post("/import/rollback")
def rollback_import(request: Request, db: Session = Depends(get_db)):
    success = restore_db_backup(db)
    
    objects_count = db.scalar(select(func.count(PropertyObject.id))) or 0
    tenants_count = db.scalar(select(func.count(Tenant.id))) or 0
    placements_count = db.scalar(select(func.count(LeasePlacement.id))) or 0
    latest_period = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).first()
    total_amount = Decimal("0.00")
    if latest_period:
        total_amount = db.scalar(
            select(func.sum(ChargeAllocation.amount))
            .where(ChargeAllocation.billing_period_id == latest_period.id)
        ) or Decimal("0.00")
        
    warnings = get_validation_warnings(db)
    msg = "Импорт успешно отменен! База данных восстановлена." if success else "Резервная копия не найдена."
    return render(
        request,
        "dashboard.html",
        db,
        objects_count=objects_count,
        tenants_count=tenants_count,
        placements_count=placements_count,
        latest_period=latest_period,
        total_amount=quantize_money(total_amount),
        success_msg=msg if success else None,
        error=None if success else msg,
        warnings=warnings,
        has_backup=False,
    )




# ==============================================================================
# PAYMENTS ENDPOINTS
# ==============================================================================

@app.get("/payments")
def payments_page(request: Request, db: Session = Depends(get_db)):
    tenants = db.scalars(select(Tenant).order_by(Tenant.display_name)).all()
    periods = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).all()
    payments = db.scalars(
        select(TenantPayment)
        .options(joinedload(TenantPayment.tenant), joinedload(TenantPayment.billing_period))
        .where(TenantPayment.is_active == True)
        .order_by(TenantPayment.payment_date.desc(), TenantPayment.id.desc())
    ).all()
    return render(
        request,
        "payments.html",
        db,
        tenants=tenants,
        periods=periods,
        payments=payments,
        form_state={},
        today=date.today().isoformat(),
    )


@app.post("/payments")
def create_payment(
    request: Request,
    tenant_id: int = Form(...),
    billing_period_id: int = Form(...),
    amount: str = Form(...),
    payment_date: date = Form(...),
    comment: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    amt = to_decimal(amount, Decimal("0.00")) or Decimal("0.00")
    if amt <= 0:
        tenants = db.scalars(select(Tenant).order_by(Tenant.display_name)).all()
        periods = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).all()
        payments = db.scalars(
            select(TenantPayment)
            .options(joinedload(TenantPayment.tenant), joinedload(TenantPayment.billing_period))
            .where(TenantPayment.is_active == True)
            .order_by(TenantPayment.payment_date.desc(), TenantPayment.id.desc())
        ).all()
        return render(
            request,
            "payments.html",
            db,
            tenants=tenants,
            periods=periods,
            payments=payments,
            error="Сумма оплаты должна быть больше нуля.",
            form_state={
                "tenant_id": tenant_id,
                "billing_period_id": billing_period_id,
                "amount": amount,
                "payment_date": payment_date.isoformat(),
                "comment": comment,
            },
            today=date.today().isoformat(),
        )

    payment = TenantPayment(
        tenant_id=tenant_id,
        billing_period_id=billing_period_id,
        amount=amt,
        payment_date=payment_date,
        comment=comment.strip() if comment else None,
    )
    db.add(payment)
    db.commit()
    return RedirectResponse("/payments", status_code=303)


@app.post("/payments/{payment_id}/edit")
def edit_payment(
    payment_id: int,
    tenant_id: int = Form(...),
    billing_period_id: int = Form(...),
    amount: str = Form(...),
    payment_date: date = Form(...),
    comment: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    payment = db.get(TenantPayment, payment_id)
    if payment is None:
        raise HTTPException(status_code=404, detail="Оплата не найдена.")

    amt = to_decimal(amount, Decimal("0.00")) or Decimal("0.00")
    if amt <= 0:
        raise HTTPException(status_code=400, detail="Сумма оплаты должна быть больше нуля.")

    payment.tenant_id = tenant_id
    payment.billing_period_id = billing_period_id
    payment.amount = amt
    payment.payment_date = payment_date
    payment.comment = comment.strip() if comment else None
    db.commit()
    return RedirectResponse("/payments", status_code=303)


@app.post("/payments/{payment_id}/delete")
def delete_payment(payment_id: int, db: Session = Depends(get_db)):
    payment = db.get(TenantPayment, payment_id)
    if payment is None:
        raise HTTPException(status_code=404, detail="Оплата не найдена.")
    
    move_to_trash(db, payment, "tenant_payment", f"Оплата арендатора {payment.tenant.display_name} на сумму {payment.amount} руб.")
    db.delete(payment)
    db.commit()
    return RedirectResponse("/payments", status_code=303)


# ==============================================================================
# SETTINGS & BACKUP & RESTORE ENDPOINTS
# ==============================================================================

@app.get("/settings")
def settings_page(request: Request, db: Session = Depends(get_db)):
    return render(request, "settings.html", db)


@app.get("/settings/backup")
def download_backup():
    from .db import DB_PATH
    if not DB_PATH.exists():
        raise HTTPException(status_code=404, detail="База данных не найдена.")
    return FileResponse(
        path=DB_PATH,
        filename=f"uk_uchet_backup_{date.today().isoformat()}.db",
        media_type="application/x-sqlite3"
    )


@app.post("/settings/restore")
def restore_backup(request: Request, backup_file: UploadFile = File(...), db: Session = Depends(get_db)):
    from .db import DB_PATH, engine
    import shutil
    import sqlite3
    import tempfile
    import os
    
    temp_fd, temp_path = tempfile.mkstemp(suffix=".db")
    os.close(temp_fd)
    
    try:
        with open(temp_path, "wb") as f:
            shutil.copyfileobj(backup_file.file, f)
        
        conn = sqlite3.connect(temp_path)
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [r[0] for r in cursor.fetchall()]
            required_tables = ["tenants", "property_objects", "lease_placements", "billing_periods"]
            for t in required_tables:
                if t not in tables:
                    raise ValueError(f"Отсутствует таблица {t}")
        except Exception as exc:
            redirect_to = request.query_params.get("redirect_to")
            if redirect_to:
                import urllib.parse
                url = f"{redirect_to}?error=" + urllib.parse.quote(f"Недопустимый файл резервной копии: {exc}")
                return RedirectResponse(url, status_code=303)
            return render(request, "settings.html", db, error=f"Недопустимый файл резервной копии: {exc}")
        finally:
            conn.close()
            
        db.close()
        engine.dispose()
        
        backup_curr = DB_PATH.with_suffix(".db.bak")
        shutil.copy2(DB_PATH, backup_curr)
        
        try:
            shutil.copy2(temp_path, DB_PATH)
            success_msg = "База данных успешно восстановлена!"
        except Exception as exc:
            shutil.copy2(backup_curr, DB_PATH)
            redirect_to = request.query_params.get("redirect_to")
            if redirect_to:
                import urllib.parse
                url = f"{redirect_to}?error=" + urllib.parse.quote(f"Ошибка восстановления: {exc}")
                return RedirectResponse(url, status_code=303)
            return render(request, "settings.html", db, error=f"Ошибка восстановления: {exc}")
        finally:
            if backup_curr.exists():
                os.remove(backup_curr)
                
        redirect_to = request.query_params.get("redirect_to")
        if redirect_to:
            import urllib.parse
            url = f"{redirect_to}?success=" + urllib.parse.quote(success_msg)
            return RedirectResponse(url, status_code=303)
        return render(request, "settings.html", db, success=success_msg)
        
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)


# ==============================================================================
# EXCEL IMPORT ENDPOINTS
# ==============================================================================

@app.get("/settings/templates/objects")
def download_objects_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Объекты"
    ws.append(["Название", "Адрес", "Общая площадь", "Примечание"])
    ws.append(["БЦ Север", "ул. Ленина, д. 10", 1500.50, "Главный офис"])
    ws.append(["ТЦ Весна", "пр. Мира, д. 25", 3400.00, "Торговый центр"])
    
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    # We need to return it as a StreamingResponse
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_objects.xlsx"}
    )


@app.get("/settings/templates/tenants")
def download_tenants_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Арендаторы"
    ws.append(["Тип (ИП/ООО)", "ФИО / Наименование", "Телефон", "Начальный баланс", "Примечание"])
    ws.append(["ИП", "Иванов Иван Иванович", "+7 (999) 111-22-33", 0.00, "Арендатор офиса 101"])
    ws.append(["ООО", "Ромашка", "+7 (999) 444-55-66", 15000.00, "Арендатор склада"])
    
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_tenants.xlsx"}
    )


@app.post("/settings/import-objects")
def import_objects(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file.file, read_only=True)
        ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Файл пуст.")
            
        header = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
        
        try:
            name_idx = header.index("название")
            address_idx = header.index("адрес")
            area_idx = header.index("общая площадь")
        except ValueError as e:
            raise ValueError(f"Отсутствует обязательная колонка (Название, Адрес, Общая площадь). Ошибка: {e}")
            
        note_idx = header.index("примечание") if "примечание" in header else -1
        
        imported_count = 0
        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue
            name = str(row[name_idx]).strip() if row[name_idx] is not None else ""
            if not name:
                continue
            
            existing = db.scalar(select(PropertyObject).where(PropertyObject.name == name))
            if existing:
                continue
                
            address = str(row[address_idx]).strip() if row[address_idx] is not None else ""
            
            try:
                area_str = str(row[area_idx]).replace(",", ".").strip()
                area = Decimal(area_str)
            except Exception:
                area = Decimal("0.00")
                
            note = str(row[note_idx]).strip() if note_idx != -1 and row[note_idx] is not None else None
            
            obj = PropertyObject(
                name=name,
                address=address,
                total_area=area,
                note=note
            )
            db.add(obj)
            imported_count += 1
            
        db.commit()
        
        redirect_to = request.query_params.get("redirect_to")
        if redirect_to:
            import urllib.parse
            url = f"{redirect_to}?success=" + urllib.parse.quote(f"Успешно импортировано объектов: {imported_count}")
            return RedirectResponse(url, status_code=303)
        return render(request, "settings.html", db, success=f"Успешно импортировано объектов: {imported_count}")
    except Exception as exc:
        redirect_to = request.query_params.get("redirect_to")
        if redirect_to:
            import urllib.parse
            url = f"{redirect_to}?error=" + urllib.parse.quote(f"Ошибка импорта: {exc}")
            return RedirectResponse(url, status_code=303)
        return render(request, "settings.html", db, error=f"Ошибка импорта: {exc}")


@app.post("/settings/import-tenants")
def import_tenants(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file.file, read_only=True)
        ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Файл пуст.")
            
        header = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
        
        try:
            type_idx = -1
            for col_name in ["тип", "тип (ип/ооо)", "тип контрагента"]:
                if col_name in header:
                    type_idx = header.index(col_name)
                    break
            if type_idx == -1:
                raise ValueError("Отсутствует колонка 'Тип (ИП/ООО)'")
                
            name_idx = -1
            for col_name in ["наименование", "фио / наименование", "имя"]:
                if col_name in header:
                    name_idx = header.index(col_name)
                    break
            if name_idx == -1:
                raise ValueError("Отсутствует колонка 'ФИО / Наименование'")
        except ValueError as e:
            raise ValueError(f"Обязательные колонки не найдены. Ошибка: {e}")
            
        phone_idx = -1
        for col_name in ["телефон", "номер телефона", "контакт"]:
            if col_name in header:
                phone_idx = header.index(col_name)
                break
                
        balance_idx = -1
        for col_name in ["начальный баланс", "баланс", "нач. баланс", "долг"]:
            if col_name in header:
                balance_idx = header.index(col_name)
                break
                
        note_idx = header.index("примечание") if "примечание" in header else -1
        
        imported_count = 0
        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue
            name = str(row[name_idx]).strip() if row[name_idx] is not None else ""
            if not name:
                continue
                
            existing = db.scalar(select(Tenant).where(Tenant.display_name == name))
            if existing:
                continue
                
            t_type = str(row[type_idx]).strip() if row[type_idx] is not None else "ИП"
            if t_type not in ["ИП", "ООО"]:
                t_type = "ИП"
                
            phone = str(row[phone_idx]).strip() if phone_idx != -1 and row[phone_idx] is not None else None
            
            try:
                bal_str = str(row[balance_idx]).replace(",", ".").strip()
                initial_balance = Decimal(bal_str)
            except Exception:
                initial_balance = Decimal("0.00")
                
            note = str(row[note_idx]).strip() if note_idx != -1 and row[note_idx] is not None else None
            
            tenant = Tenant(
                tenant_type=t_type,
                display_name=name,
                phone=phone,
                initial_balance=initial_balance,
                note=note
            )
            db.add(tenant)
            imported_count += 1
            
        db.commit()
        
        redirect_to = request.query_params.get("redirect_to")
        if redirect_to:
            import urllib.parse
            url = f"{redirect_to}?success=" + urllib.parse.quote(f"Успешно импортировано арендаторов: {imported_count}")
            return RedirectResponse(url, status_code=303)
        return render(request, "settings.html", db, success=f"Успешно импортировано арендаторов: {imported_count}")
    except Exception as exc:
        redirect_to = request.query_params.get("redirect_to")
        if redirect_to:
            import urllib.parse
            url = f"{redirect_to}?error=" + urllib.parse.quote(f"Ошибка импорта: {exc}")
            return RedirectResponse(url, status_code=303)
        return render(request, "settings.html", db, error=f"Ошибка импорта: {exc}")


# =========================================================================
# V2 ROUTING (SIMPLIFIED WIZARD INTERFACE)
# =========================================================================

@app.get("/v2/directory")
def v2_directory_page(request: Request, db: Session = Depends(get_db)):
    objects = db.scalars(select(PropertyObject).order_by(PropertyObject.name)).all()
    occupied_areas: dict[int, Decimal] = {}
    for obj in objects:
        active_placements = db.scalars(
            select(LeasePlacement).where(
                LeasePlacement.object_id == obj.id,
                LeasePlacement.is_active.is_(True)
            )
        ).all()
        occupied_areas[obj.id] = sum(Decimal(p.occupied_area) for p in active_placements)
        
    tenants = db.scalars(select(Tenant).order_by(Tenant.display_name)).all()
    
    placements = db.scalars(
        select(LeasePlacement)
        .options(joinedload(LeasePlacement.object), joinedload(LeasePlacement.tenant))
        .order_by(LeasePlacement.start_date.desc())
    ).all()
    
    return render(
        request,
        "v2/directory.html",
        db,
        objects=objects,
        occupied_areas=occupied_areas,
        tenants=tenants,
        placements=placements,
        error=request.query_params.get("error"),
        success=request.query_params.get("success"),
        form_state={},
    )


@app.get("/v2/billing")
def v2_billing_page(request: Request, db: Session = Depends(get_db)):
    periods = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).all()
    selected_id = request.query_params.get("period_id")
    selected_period = None
    if selected_id and selected_id.isdigit():
        selected_period = db.get(BillingPeriod, int(selected_id))
    if not selected_period and periods:
        selected_period = periods[0]
    
    allocations = []
    totals: dict[str, Decimal] = {}
    tenant_balances = {}
    tenant_objects = {}
    charges = []
    
    if selected_period:
        allocations = db.scalars(
            select(ChargeAllocation)
            .options(
                joinedload(ChargeAllocation.object),
                joinedload(ChargeAllocation.tenant),
                joinedload(ChargeAllocation.utility_charge),
            )
            .where(ChargeAllocation.billing_period_id == selected_period.id)
            .order_by(ChargeAllocation.object_id, ChargeAllocation.tenant_id)
        ).all()
        tenant_balances = get_tenant_balances(db, selected_period)
        tenant_objects = {t.id: t for t in db.scalars(select(Tenant)).all()}
        for tenant_id, bal in tenant_balances.items():
            t = tenant_objects[tenant_id]
            totals[t.display_name] = bal["allocated"]
            
        charges = db.scalars(
            select(UtilityCharge)
            .options(joinedload(UtilityCharge.object))
            .where(UtilityCharge.billing_period_id == selected_period.id)
            .order_by(UtilityCharge.utility_type)
        ).all()
            
    objects = db.scalars(select(PropertyObject).order_by(PropertyObject.name)).all()
    tenants = db.scalars(select(Tenant).order_by(Tenant.display_name)).all()
    grouped_periods = get_grouped_documents(db)
    
    warnings = get_validation_warnings(db)

    return render(
        request,
        "v2/billing.html",
        db,
        periods=periods,
        selected_period=selected_period,
        allocations=allocations,
        totals={k: quantize_money(v) for k, v in totals.items()},
        tenant_balances=tenant_balances,
        tenant_objects=tenant_objects,
        charges=charges,
        objects=objects,
        tenants=tenants,
        grouped_periods=grouped_periods,
        warnings=warnings,
        error=request.query_params.get("error"),
        success=request.query_params.get("success"),
    )


@app.get("/v2/payments")
def v2_payments_page(request: Request, db: Session = Depends(get_db)):
    payments = db.scalars(
        select(TenantPayment)
        .options(joinedload(TenantPayment.tenant), joinedload(TenantPayment.billing_period))
        .order_by(TenantPayment.payment_date.desc(), TenantPayment.created_at.desc())
    ).all()
    tenants = db.scalars(select(Tenant).order_by(Tenant.display_name)).all()
    periods = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).all()
    
    return render(
        request,
        "v2/payments.html",
        db,
        payments=payments,
        tenants=tenants,
        periods=periods,
        error=request.query_params.get("error"),
        success=request.query_params.get("success"),
        form_state={},
    )


@app.get("/v2/settings")
def v2_settings_page(request: Request, db: Session = Depends(get_db)):
    from .paths import DATA_DIR
    has_backup = (DATA_DIR / "import_backup.json").exists()
    trash_items = db.scalars(select(TrashBin).order_by(TrashBin.deleted_at.desc())).all()
    
    tariffs = db.scalars(
        select(Tariff)
        .options(joinedload(Tariff.object), joinedload(Tariff.tenant))
        .order_by(Tariff.utility_type, Tariff.start_date.desc())
    ).all()
    
    rules = db.scalars(
        select(AllocationRule)
        .options(joinedload(AllocationRule.object), joinedload(AllocationRule.tenant))
        .order_by(AllocationRule.utility_type)
    ).all()
    
    objects = db.scalars(select(PropertyObject).order_by(PropertyObject.name)).all()
    tenants = db.scalars(select(Tenant).order_by(Tenant.display_name)).all()
    
    return render(
        request,
        "v2/settings.html",
        db,
        has_backup=has_backup,
        trash_items=trash_items,
        tariffs=tariffs,
        rules=rules,
        objects=objects,
        tenants=tenants,
        error=request.query_params.get("error"),
        success=request.query_params.get("success"),
    )



@app.get('/v2/guide')
def v2_guide_page(request: Request, db: Session = Depends(get_db)):
    return render(request, 'v2/guide.html', db)



@app.get("/")
def landing_page(request: Request, db: Session = Depends(get_db)):
    return render(request, "landing.html", db)
