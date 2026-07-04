from __future__ import annotations

from collections import defaultdict
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import List, Optional, Dict
import zipfile
import tempfile

from fastapi import Depends, FastAPI, Form, HTTPException, Request, UploadFile, File
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import io
from sqlalchemy.exc import OperationalError
from sqlalchemy import func, select, delete, text
from sqlalchemy.orm import Session, joinedload
from .db import Base, SessionLocal, engine, get_db
from .models import (
    AllocationRule,
    BillingPeriod,
    ChargeAllocation,
    GeneratedDocument,
    LeasePlacement,
    PropertyObject,
    Tenant,
    UtilityCharge,
    TrashBin,
    Tariff,
)
from .paths import BUNDLE_ROOT
from .services.calculations import close_period, get_charge_amount, quantize_money, recalculate_period, overlaps
from .services.documents import (
    generate_act_docx,
    generate_invoice_docx,
    generate_register_xlsx,
    generate_invoice_xlsx,
    generate_act_xlsx,
    _safe_name,
)



try:
    Base.metadata.create_all(bind=engine, checkfirst=True)
except OperationalError as exc:
    if "already exists" not in str(exc):
        raise

app = FastAPI(title="Utility Billing App")
RESOURCE_DIR = BUNDLE_ROOT / "app"
templates = Jinja2Templates(directory=str(RESOURCE_DIR / "templates"))
app.mount("/static", StaticFiles(directory=str(RESOURCE_DIR / "static")), name="static")

@app.get("/favicon.ico", include_in_schema=False)
def favicon():
    ico_path = BUNDLE_ROOT / "app_icon.ico"
    if ico_path.exists():
        return FileResponse(ico_path)
    fallback_path = RESOURCE_DIR / "static" / "app_icon.png"
    if fallback_path.exists():
        return FileResponse(fallback_path)
    raise HTTPException(status_code=404)

MAX_OBJECTS = 10
MAX_TENANTS = 50
UTILITY_TYPES = ("heat", "electricity", "water", "cleaning")
TENANT_TYPES = ("ИП", "ООО")
UTILITY_LABELS = {
    "heat": "Тепло",
    "electricity": "Электричество",
    "water": "Вода",
    "cleaning": "Уборка",
}


@app.get("/__health")
def health_check():
    return {"app": "RealEstateUtilityApp"}


def to_decimal(value: Optional[str], default: Optional[Decimal] = None) -> Optional[Decimal]:
    if value in (None, ""):
        return default
    try:
        return Decimal(value)
    except InvalidOperation as exc:
        raise HTTPException(status_code=400, detail=f"Некорректное число: {value}") from exc


def require_choice(value: str, allowed: tuple[str, ...], field_name: str) -> None:
    if value not in allowed:
        raise HTTPException(status_code=400, detail=f"Недопустимое значение поля '{field_name}'.")


def require_positive(value: Decimal, field_name: str) -> None:
    if value < 0:
        raise HTTPException(status_code=400, detail=f"Поле '{field_name}' не может быть отрицательным.")


STATUS_LABELS = {
    "draft": "Черновик",
    "calculated": "Рассчитан",
    "closed": "Закрыт",
}

PERIOD_TYPE_LABELS = {
    "month": "Месяц",
    "range": "Диапазон",
}

ALLOCATION_MODE_LABELS = {
    "area": "По площади",
    "manual": "Вручную",
    "mixed": "Смешанное",
}

BASE_AREA_MODE_LABELS = {
    "active_leases": "Сумма активных площадей",
    "object_total": "Общая площадь объекта",
}

VALUE_TYPE_LABELS = {
    "percent": "Процент",
    "fixed": "Фиксированная сумма",
}


def render(request: Request, page: str, db: Session, status_code: int = 200, **context):
    summary = {
        "objects": db.scalar(select(func.count(PropertyObject.id))) or 0,
        "tenants": db.scalar(select(func.count(Tenant.id))) or 0,
        "placements": db.scalar(select(func.count(LeasePlacement.id))) or 0,
        "periods": db.scalar(select(func.count(BillingPeriod.id))) or 0,
    }
    base_context = {
        "request": request,
        "page": page,
        "summary": summary,
        "utility_labels": UTILITY_LABELS,
        "status_labels": STATUS_LABELS,
        "period_type_labels": PERIOD_TYPE_LABELS,
        "allocation_mode_labels": ALLOCATION_MODE_LABELS,
        "base_area_mode_labels": BASE_AREA_MODE_LABELS,
        "value_type_labels": VALUE_TYPE_LABELS,
        "today": date.today(),
    }
    base_context.update(context)
    return templates.TemplateResponse(request, page, base_context, status_code=status_code)


def render_billing_page(
    request: Request,
    db: Session,
    *,
    selected_period_id: Optional[int] = None,
    error: Optional[str] = None,
    status_code: int = 200,
):
    periods = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).all()
    if selected_period_id is not None:
        selected_period = db.get(BillingPeriod, selected_period_id)
    else:
        selected_id = request.query_params.get("period_id")
        selected_period = db.get(BillingPeriod, int(selected_id)) if selected_id else (periods[0] if periods else None)
    allocations = []
    totals: dict[str, Decimal] = {}
    if selected_period:
        allocations = db.scalars(
            select(ChargeAllocation)
            .options(
                joinedload(ChargeAllocation.object),
                joinedload(ChargeAllocation.tenant),
                joinedload(ChargeAllocation.utility_charge),
            )
            .where(ChargeAllocation.billing_period_id == selected_period.id)
            .order_by(ChargeAllocation.object_id, ChargeAllocation.tenant_id)
        ).all()
        for allocation in allocations:
            totals.setdefault(allocation.tenant.display_name, Decimal("0.00"))
            totals[allocation.tenant.display_name] += Decimal(allocation.amount)
    return render(
        request,
        "billing.html",
        db,
        status_code=status_code,
        periods=periods,
        selected_period=selected_period,
        allocations=allocations,
        totals={k: quantize_money(v) for k, v in totals.items()},
        error=error,
    )


def get_grouped_documents(db: Session):
    periods = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).all()
    documents = db.scalars(
        select(GeneratedDocument)
        .options(joinedload(GeneratedDocument.tenant), joinedload(GeneratedDocument.billing_period))
        .order_by(GeneratedDocument.created_at.desc())
    ).all()
    
    period_map = {}
    for period in periods:
        period_map[period.id] = {
            "period": period,
            "registers": [],
            "tenants": {}
        }
        
    for doc in documents:
        p_id = doc.billing_period_id
        if p_id not in period_map:
            continue
        
        if doc.document_type == "register" and doc.tenant_id is None:
            if not period_map[p_id]["registers"]:
                period_map[p_id]["registers"].append(doc)
        elif doc.tenant:
            t_id = doc.tenant_id
            t_map = period_map[p_id]["tenants"]
            if t_id not in t_map:
                t_map[t_id] = {
                    "tenant_name": doc.tenant.display_name,
                    "register": None,
                    "invoice_docx": None,
                    "invoice_xlsx": None,
                    "act_docx": None,
                    "act_xlsx": None
                }
            if doc.document_type == "register":
                if t_map[t_id]["register"] is None:
                    t_map[t_id]["register"] = doc
            elif doc.document_type == "invoice":
                if t_map[t_id]["invoice_docx"] is None:
                    t_map[t_id]["invoice_docx"] = doc
            elif doc.document_type == "invoice_xlsx":
                if t_map[t_id]["invoice_xlsx"] is None:
                    t_map[t_id]["invoice_xlsx"] = doc
            elif doc.document_type == "act":
                if t_map[t_id]["act_docx"] is None:
                    t_map[t_id]["act_docx"] = doc
            elif doc.document_type == "act_xlsx":
                if t_map[t_id]["act_xlsx"] is None:
                    t_map[t_id]["act_xlsx"] = doc
                
    result = []
    for p_id in sorted(period_map.keys(), key=lambda k: period_map[k]["period"].start_date, reverse=True):
        p_data = period_map[p_id]
        sorted_tenants = sorted(p_data["tenants"].values(), key=lambda x: x["tenant_name"])
        result.append({
            "period": p_data["period"],
            "registers": p_data["registers"],
            "tenants": sorted_tenants
        })
    return result


def render_documents_page(
    request: Request,
    db: Session,
    *,
    error: Optional[str] = None,
    form_state: Optional[dict[str, object]] = None,
    status_code: int = 200,
):
    periods = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).all()
    tenants = db.scalars(select(Tenant).order_by(Tenant.display_name)).all()
    grouped_periods = get_grouped_documents(db)
    return render(
        request,
        "documents.html",
        db,
        status_code=status_code,
        periods=periods,
        tenants=tenants,
        grouped_periods=grouped_periods,
        error=error,
        form_state=form_state or {},
    )


@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    if request.url.path == "/placements":
        form = await request.form()

        def _as_int(value):
            try:
                return int(value)
            except (TypeError, ValueError):
                return value

        form_state = {
            "object_id": _as_int(form.get("object_id")),
            "tenant_id": _as_int(form.get("tenant_id")),
            "rental_address": form.get("rental_address", ""),
            "occupied_area": form.get("occupied_area", ""),
            "start_date": form.get("start_date", ""),
            "end_date": form.get("end_date", ""),
            "is_active": form.get("is_active") in ("on", "true", "True", "1"),
            "rent_tariff": form.get("rent_tariff", ""),
            "status": form.get("status", "active"),
            "comment": form.get("comment", ""),
        }
        db = SessionLocal()
        try:
            return render_placements_page(
                request,
                db,
                error=str(exc.detail),
                form_state=form_state,
                status_code=exc.status_code,
            )
        finally:
            db.close()
    return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})


def recalculate_drafts(db: Session):
    draft_periods = db.scalars(select(BillingPeriod).where(BillingPeriod.status == "draft")).all()
    for p in draft_periods:
        try:
            recalculate_period(db, p.id)
        except Exception:
            try:
                db.execute(delete(ChargeAllocation).where(ChargeAllocation.billing_period_id == p.id))
                p.status = "draft"
                db.commit()
            except Exception:
                db.rollback()


def make_db_backup(db: Session):
    backup = {
        "objects": [{k: v for k, v in obj.__dict__.items() if k != "_sa_instance_state"} for obj in db.scalars(select(PropertyObject)).all()],
        "tenants": [{k: v for k, v in t.__dict__.items() if k != "_sa_instance_state"} for t in db.scalars(select(Tenant)).all()],
        "placements": [{k: v for k, v in p.__dict__.items() if k != "_sa_instance_state"} for p in db.scalars(select(LeasePlacement)).all()],
        "rules": [{k: v for k, v in r.__dict__.items() if k != "_sa_instance_state"} for r in db.scalars(select(AllocationRule)).all()],
        "charges": [{k: v for k, v in c.__dict__.items() if k != "_sa_instance_state"} for c in db.scalars(select(UtilityCharge)).all()],
        "allocations": [{k: v for k, v in a.__dict__.items() if k != "_sa_instance_state"} for a in db.scalars(select(ChargeAllocation)).all()],
        "periods": [{k: v for k, v in p.__dict__.items() if k != "_sa_instance_state"} for p in db.scalars(select(BillingPeriod)).all()]
    }
    from .paths import DATA_DIR
    backup_file = DATA_DIR / "import_backup.json"
    with open(backup_file, "w", encoding="utf-8") as f:
        json.dump(backup, f, cls=DbEncoder, ensure_ascii=False)


def restore_db_backup(db: Session) -> bool:
    from .paths import DATA_DIR
    backup_file = DATA_DIR / "import_backup.json"
    if not backup_file.exists():
        return False
        
    with open(backup_file, "r", encoding="utf-8") as f:
        backup = json.load(f)
        
    # Truncate all tables
    db.execute(delete(ChargeAllocation))
    db.execute(delete(UtilityCharge))
    db.execute(delete(AllocationRule))
    db.execute(delete(LeasePlacement))
    db.execute(delete(Tenant))
    db.execute(delete(PropertyObject))
    db.execute(delete(BillingPeriod))
    db.flush()
    
    # Restore
    for data in backup.get("objects", []):
        db.add(PropertyObject(
            id=data.get("id"),
            name=data.get("name"),
            address=data.get("address"),
            total_area=Decimal(data.get("total_area")),
            note=data.get("note")
        ))
    for data in backup.get("tenants", []):
        db.add(Tenant(
            id=data.get("id"),
            tenant_type=data.get("tenant_type"),
            display_name=data.get("display_name"),
            phone=data.get("phone"),
            note=data.get("note")
        ))
    db.flush()
    
    for data in backup.get("periods", []):
        db.add(BillingPeriod(
            id=data.get("id"),
            period_type=data.get("period_type"),
            month_label=data.get("month_label"),
            start_date=date.fromisoformat(data.get("start_date")),
            end_date=date.fromisoformat(data.get("end_date")),
            status=data.get("status")
        ))
    db.flush()
    
    for data in backup.get("placements", []):
        db.add(LeasePlacement(
            id=data.get("id"),
            object_id=data.get("object_id"),
            tenant_id=data.get("tenant_id"),
            rental_address=data.get("rental_address"),
            occupied_area=Decimal(data.get("occupied_area")),
            start_date=date.fromisoformat(data.get("start_date")),
            end_date=date.fromisoformat(data.get("end_date")) if data.get("end_date") else None,
            is_active=data.get("is_active")
        ))
    for data in backup.get("rules", []):
        db.add(AllocationRule(
            id=data.get("id"),
            object_id=data.get("object_id"),
            utility_type=data.get("utility_type"),
            mode=data.get("mode"),
            base_area_mode=data.get("base_area_mode"),
            tenant_id=data.get("tenant_id"),
            value_type=data.get("value_type"),
            value=Decimal(data.get("value")) if data.get("value") else None,
            is_active=data.get("is_active")
        ))
    for data in backup.get("charges", []):
        db.add(UtilityCharge(
            id=data.get("id"),
            object_id=data.get("object_id"),
            billing_period_id=data.get("billing_period_id"),
            utility_type=data.get("utility_type"),
            input_mode=data.get("input_mode"),
            allocation_mode=data.get("allocation_mode"),
            amount=Decimal(data.get("amount")) if data.get("amount") else None,
            meter_from=Decimal(data.get("meter_from")) if data.get("meter_from") else None,
            meter_to=Decimal(data.get("meter_to")) if data.get("meter_to") else None,
            tariff=Decimal(data.get("tariff")) if data.get("tariff") else None,
            unit_name=data.get("unit_name"),
            comment=data.get("comment")
        ))
    db.flush()
    
    for data in backup.get("allocations", []):
        db.add(ChargeAllocation(
            id=data.get("id"),
            billing_period_id=data.get("billing_period_id"),
            utility_charge_id=data.get("utility_charge_id"),
            object_id=data.get("object_id"),
            tenant_id=data.get("tenant_id"),
            base_area=Decimal(data.get("base_area")),
            share_value=Decimal(data.get("share_value")),
            amount=Decimal(data.get("amount")),
            mode=data.get("mode"),
            manual_override=data.get("manual_override"),
            audit_payload=data.get("audit_payload")
        ))
        
    db.commit()
    
    try:
        backup_file.unlink()
    except Exception:
        pass
    return True


def get_validation_warnings(db: Session) -> List[str]:
    warnings = []
    
    # 1. Sum of tenant areas exceeds total area for an object
    objects = db.scalars(select(PropertyObject)).all()
    for obj in objects:
        active_placements = db.scalars(
            select(LeasePlacement)
            .where(LeasePlacement.object_id == obj.id, LeasePlacement.is_active.is_(True))
        ).all()
        total_occupied = sum(p.occupied_area for p in active_placements)
        if total_occupied > obj.total_area:
            warnings.append(
                f"Сумма активных площадей арендаторов в объекте \"{obj.name}\" ({total_occupied} кв.м) превышает общую площадь здания ({obj.total_area} кв.м)."
            )
            
    # Get latest period to use for "current period" checks
    latest_period = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).first()
    if latest_period:
        # 2. Tenant has no active placements in the latest period
        tenants = db.scalars(select(Tenant)).all()
        for tenant in tenants:
            active_in_period = db.scalars(
                select(LeasePlacement)
                .where(
                    LeasePlacement.tenant_id == tenant.id,
                    LeasePlacement.is_active.is_(True)
                )
            ).all()
            overlapping = [
                p for p in active_in_period 
                if overlaps(p.start_date, p.end_date, latest_period.start_date, latest_period.end_date)
            ]
            if not overlapping:
                warnings.append(
                    f"У арендатора \"{tenant.display_name}\" нет активных договоров размещения в текущем периоде ({latest_period.month_label or (latest_period.start_date.isoformat() + ' - ' + latest_period.end_date.isoformat())})."
                )
                
        # 3. Object has a utility charge in the latest period but no rule configured
        charges = db.scalars(
            select(UtilityCharge)
            .options(joinedload(UtilityCharge.object))
            .where(UtilityCharge.billing_period_id == latest_period.id)
        ).all()
        warned = set()
        for charge in charges:
            key = (charge.object_id, charge.utility_type)
            if key in warned:
                continue
            has_rule = db.scalar(
                select(func.count(AllocationRule.id))
                .where(
                    AllocationRule.object_id == charge.object_id,
                    AllocationRule.utility_type == charge.utility_type,
                    AllocationRule.is_active.is_(True)
                )
            )
            if not has_rule:
                utility_name = UTILITY_LABELS.get(charge.utility_type, charge.utility_type)
                warnings.append(
                    f"Для объекта \"{charge.object.name}\" внесен счет за \"{utility_name}\" в текущем периоде, но не настроено ни одно правило распределения."
                )
                warned.add(key)
    return warnings


import json
from datetime import datetime, date
from decimal import Decimal

class DbEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        if isinstance(obj, Decimal):
            return str(obj)
        return super().default(obj)

def move_to_trash(db: Session, record, entity_type: str, display_name: str):
    data = {k: v for k, v in record.__dict__.items() if k != "_sa_instance_state"}
    trash_item = TrashBin(
        entity_type=entity_type,
        original_id=record.id,
        display_name=display_name,
        data_json=json.dumps(data, cls=DbEncoder)
    )
    db.add(trash_item)
    db.flush()


@app.get("/")
def dashboard_page(request: Request, db: Session = Depends(get_db)):
    objects_count = db.scalar(select(func.count(PropertyObject.id))) or 0
    tenants_count = db.scalar(select(func.count(Tenant.id))) or 0
    placements_count = db.scalar(select(func.count(LeasePlacement.id))) or 0
    
    latest_period = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).first()
    total_amount = Decimal("0.00")
    if latest_period:
        total_amount = db.scalar(
            select(func.sum(ChargeAllocation.amount))
            .where(ChargeAllocation.billing_period_id == latest_period.id)
        ) or Decimal("0.00")
        
    warnings = get_validation_warnings(db)
    from .paths import DATA_DIR
    has_backup = (DATA_DIR / "import_backup.json").exists()
        
    return render(
        request,
        "dashboard.html",
        db,
        objects_count=objects_count,
        tenants_count=tenants_count,
        placements_count=placements_count,
        latest_period=latest_period,
        total_amount=quantize_money(total_amount),
        warnings=warnings,
        has_backup=has_backup,
    )



@app.get("/objects")
def objects_page(request: Request, db: Session = Depends(get_db)):
    objects = db.scalars(select(PropertyObject).order_by(PropertyObject.name)).all()
    # Рассчитываем занятую площадь по каждому объекту
    occupied_areas: dict[int, Decimal] = {}
    for obj in objects:
        active_placements = db.scalars(
            select(LeasePlacement).where(
                LeasePlacement.object_id == obj.id,
                LeasePlacement.is_active.is_(True)
            )
        ).all()
        occupied_areas[obj.id] = sum(
            Decimal(p.occupied_area) for p in active_placements
        )
    return render(request, "objects.html", db, objects=objects, occupied_areas=occupied_areas)


@app.post("/objects")
def create_object(
    name: str = Form(...),
    address: str = Form(""),
    total_area: str = Form("0.00"),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    if (db.scalar(select(func.count(PropertyObject.id))) or 0) >= MAX_OBJECTS:
        raise HTTPException(status_code=400, detail="Нельзя добавить больше 10 объектов.")

    total_area_value = to_decimal(total_area, Decimal("0.00")) or Decimal("0.00")
    require_positive(total_area_value, "общая площадь")

    obj = PropertyObject(
        name=name.strip(),
        address=address.strip() if address and address.strip() else "Не указан",
        total_area=total_area_value,
        note=note.strip() or None,
    )
    db.add(obj)
    db.commit()
    return RedirectResponse("/objects", status_code=303)


@app.post("/objects/{object_id}/edit")
def edit_object(
    object_id: int,
    name: str = Form(...),
    address: str = Form(""),
    total_area: str = Form("0.00"),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    obj = db.get(PropertyObject, object_id)
    if obj is None:
        raise HTTPException(status_code=404, detail="Объект не найден.")
    total_area_value = to_decimal(total_area, Decimal("0.00")) or Decimal("0.00")
    require_positive(total_area_value, "общая площадь")
    obj.name = name.strip()
    obj.address = address.strip() if address and address.strip() else "Не указан"
    obj.total_area = total_area_value
    obj.note = note.strip() or None
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/objects", status_code=303)


@app.post("/objects/{object_id}/delete")
def delete_object(object_id: int, db: Session = Depends(get_db)):
    obj = db.get(PropertyObject, object_id)
    if obj is None:
        raise HTTPException(status_code=404, detail="Объект не найден.")
    move_to_trash(db, obj, "property_object", f"Объект: {obj.name} ({obj.total_area} кв.м)")
    db.execute(delete(ChargeAllocation).where(ChargeAllocation.object_id == object_id))
    db.execute(delete(UtilityCharge).where(UtilityCharge.object_id == object_id))
    db.execute(delete(LeasePlacement).where(LeasePlacement.object_id == object_id))
    db.execute(delete(AllocationRule).where(AllocationRule.object_id == object_id))
    db.delete(obj)
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/objects", status_code=303)


@app.get("/tenants")
def tenants_page(request: Request, db: Session = Depends(get_db)):
    tenants = db.scalars(select(Tenant).order_by(Tenant.display_name)).all()
    return render(request, "tenants.html", db, tenants=tenants)


@app.post("/tenants")
def create_tenant(
    tenant_type: str = Form(...),
    display_name: str = Form(...),
    phone: str = Form(""),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    if (db.scalar(select(func.count(Tenant.id))) or 0) >= MAX_TENANTS:
        raise HTTPException(status_code=400, detail="Нельзя добавить больше 50 арендаторов.")

    require_choice(tenant_type, TENANT_TYPES, "тип арендатора")

    tenant = Tenant(
        tenant_type=tenant_type,
        display_name=display_name.strip(),
        phone=phone.strip() or None,
        note=note.strip() or None,
    )
    db.add(tenant)
    db.commit()
    return RedirectResponse("/tenants", status_code=303)


@app.post("/tenants/{tenant_id}/edit")
def edit_tenant(
    tenant_id: int,
    tenant_type: str = Form(...),
    display_name: str = Form(...),
    phone: str = Form(""),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    tenant = db.get(Tenant, tenant_id)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Арендатор не найден.")
    require_choice(tenant_type, TENANT_TYPES, "тип арендатора")
    tenant.tenant_type = tenant_type
    tenant.display_name = display_name.strip()
    tenant.phone = phone.strip() or None
    tenant.note = note.strip() or None
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/tenants", status_code=303)


@app.post("/tenants/{tenant_id}/delete")
def delete_tenant(tenant_id: int, db: Session = Depends(get_db)):
    tenant = db.get(Tenant, tenant_id)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Арендатор не найден.")
    move_to_trash(db, tenant, "tenant", f"Арендатор: {tenant.tenant_type} {tenant.display_name}")
    db.execute(delete(ChargeAllocation).where(ChargeAllocation.tenant_id == tenant_id))
    db.execute(delete(LeasePlacement).where(LeasePlacement.tenant_id == tenant_id))
    db.execute(delete(AllocationRule).where(AllocationRule.tenant_id == tenant_id))
    db.execute(delete(GeneratedDocument).where(GeneratedDocument.tenant_id == tenant_id))
    db.delete(tenant)
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/tenants", status_code=303)


@app.get("/placements")
def placements_page(request: Request, db: Session = Depends(get_db)):
    return render_placements_page(request, db)


def render_placements_page(
    request: Request,
    db: Session,
    *,
    error: Optional[str] = None,
    form_state: Optional[dict[str, object]] = None,
    status_code: int = 200,
):
    placements = db.scalars(
        select(LeasePlacement)
        .options(joinedload(LeasePlacement.object), joinedload(LeasePlacement.tenant))
        .order_by(LeasePlacement.start_date.desc())
    ).all()
    objects = db.scalars(select(PropertyObject).order_by(PropertyObject.name)).all()
    tenants = db.scalars(select(Tenant).order_by(Tenant.display_name)).all()
    return render(
        request,
        "placements.html",
        db,
        status_code=status_code,
        placements=placements,
        objects=objects,
        tenants=tenants,
        error=error,
        form_state=form_state or {},
    )


@app.post("/placements")
def create_placement(
    request: Request,
    object_id: int = Form(...),
    tenant_id: int = Form(...),
    rental_address: str = Form(...),
    occupied_area: str = Form(...),
    start_date: date = Form(...),
    end_date: Optional[date] = Form(None),
    is_active: bool = Form(False),
    rent_tariff: Optional[str] = Form(None),
    status: str = Form("active"),
    comment: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    obj = db.get(PropertyObject, object_id)
    if obj is None:
        raise HTTPException(status_code=404, detail="Объект не найден.")
    if db.get(Tenant, tenant_id) is None:
        raise HTTPException(status_code=404, detail="Арендатор не найден.")

    area = to_decimal(occupied_area, Decimal("0")) or Decimal("0")
    require_positive(area, "занимаемая площадь")
    
    rent_tariff_val = to_decimal(rent_tariff, None)
    if rent_tariff_val is not None:
        require_positive(rent_tariff_val, "тариф аренды")

    if area > Decimal(obj.total_area):
        return render_placements_page(
            request,
            db,
            error=f"Занимаемая площадь ({area}) не может быть больше общей площади объекта ({obj.total_area}).",
            form_state={
                "object_id": object_id,
                "tenant_id": tenant_id,
                "rental_address": rental_address,
                "occupied_area": occupied_area,
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat() if end_date else "",
                "is_active": is_active,
                "rent_tariff": rent_tariff,
                "status": status,
                "comment": comment,
            },
            status_code=400,
        )

    placement = LeasePlacement(
        object_id=object_id,
        tenant_id=tenant_id,
        rental_address=rental_address.strip(),
        occupied_area=area,
        start_date=start_date,
        end_date=end_date,
        is_active=is_active,
        rent_tariff=rent_tariff_val,
        status=status.strip(),
        comment=comment.strip() if comment else None,
    )
    db.add(placement)
    db.commit()
    # Recalculate draft periods
    draft_periods = db.scalars(select(BillingPeriod).where(BillingPeriod.status == "draft")).all()
    for p in draft_periods:
        try:
            recalculate_period(db, p.id)
        except Exception:
            pass
    return RedirectResponse("/placements", status_code=303)


@app.post("/placements/{placement_id}/edit")
def edit_placement(
    request: Request,
    placement_id: int,
    object_id: int = Form(...),
    tenant_id: int = Form(...),
    rental_address: str = Form(...),
    occupied_area: str = Form(...),
    start_date: date = Form(...),
    end_date: Optional[date] = Form(None),
    is_active: bool = Form(False),
    rent_tariff: Optional[str] = Form(None),
    status: str = Form("active"),
    comment: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    placement = db.get(LeasePlacement, placement_id)
    if placement is None:
        raise HTTPException(status_code=404, detail="Размещение не найдено.")
    obj = db.get(PropertyObject, object_id)
    if obj is None:
        raise HTTPException(status_code=404, detail="Объект не найден.")
    if db.get(Tenant, tenant_id) is None:
        raise HTTPException(status_code=404, detail="Арендатор не найден.")

    area = to_decimal(occupied_area, Decimal("0")) or Decimal("0")
    require_positive(area, "занимаемая площадь")
    
    rent_tariff_val = to_decimal(rent_tariff, None)
    if rent_tariff_val is not None:
        require_positive(rent_tariff_val, "тариф аренды")

    if area > Decimal(obj.total_area):
        return render_placements_page(
            request,
            db,
            error=f"Занимаемая площадь ({area}) не может быть больше общей площади объекта ({obj.total_area}).",
            status_code=400,
        )

    placement.object_id = object_id
    placement.tenant_id = tenant_id
    placement.rental_address = rental_address.strip()
    placement.occupied_area = area
    placement.start_date = start_date
    placement.end_date = end_date
    placement.is_active = is_active
    placement.rent_tariff = rent_tariff_val
    placement.status = status.strip()
    placement.comment = comment.strip() if comment else None
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/placements", status_code=303)


@app.post("/placements/{placement_id}/delete")
def delete_placement(placement_id: int, db: Session = Depends(get_db)):
    placement = db.get(LeasePlacement, placement_id)
    if placement is None:
        raise HTTPException(status_code=404, detail="Размещение не найдено.")
    move_to_trash(db, placement, "lease_placement", f"Размещение: {placement.tenant.display_name} -> {placement.object.name} ({placement.occupied_area} кв.м)")
    db.delete(placement)
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/placements", status_code=303)


@app.get("/charges")
def charges_page(request: Request, db: Session = Depends(get_db)):
    objects = db.scalars(select(PropertyObject).order_by(PropertyObject.name)).all()
    periods = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).all()
    tenants = db.scalars(select(Tenant).order_by(Tenant.display_name)).all()
    charges = db.scalars(
        select(UtilityCharge)
        .options(joinedload(UtilityCharge.object), joinedload(UtilityCharge.billing_period))
        .order_by(UtilityCharge.created_at.desc())
    ).all()
    rules = db.scalars(
        select(AllocationRule)
        .options(joinedload(AllocationRule.object), joinedload(AllocationRule.tenant))
        .order_by(AllocationRule.created_at.desc())
    ).all()
    tariffs = db.scalars(
        select(Tariff)
        .options(joinedload(Tariff.object), joinedload(Tariff.tenant))
        .order_by(Tariff.created_at.desc())
    ).all()
    return render(
        request,
        "charges.html",
        db,
        objects=objects,
        periods=periods,
        tenants=tenants,
        charges=charges,
        rules=rules,
        tariffs=tariffs,
    )


@app.post("/charges")
def create_charge(
    object_id: int = Form(...),
    billing_period_id: int = Form(...),
    utility_type: str = Form(...),
    input_mode: str = Form(...),
    allocation_mode: str = Form(...),
    amount: str = Form(""),
    meter_from: str = Form(""),
    meter_to: str = Form(""),
    tariff: str = Form(""),
    unit_name: str = Form(""),
    comment: str = Form(""),
    db: Session = Depends(get_db),
):
    if db.get(PropertyObject, object_id) is None:
        raise HTTPException(status_code=404, detail="Объект не найден.")
    if db.get(BillingPeriod, billing_period_id) is None:
        raise HTTPException(status_code=404, detail="Период не найден.")
    require_choice(utility_type, UTILITY_TYPES, "услуга")

    charge = UtilityCharge(
        object_id=object_id,
        billing_period_id=billing_period_id,
        utility_type=utility_type,
        input_mode=input_mode,
        allocation_mode=allocation_mode,
        amount=to_decimal(amount),
        meter_from=to_decimal(meter_from),
        meter_to=to_decimal(meter_to),
        tariff=to_decimal(tariff),
        unit_name=unit_name.strip() or None,
        comment=comment.strip() or None,
    )
    get_charge_amount(charge)
    db.add(charge)
    db.commit()
    try:
        recalculate_period(db, billing_period_id)
    except Exception:
        pass
    return RedirectResponse("/charges", status_code=303)


@app.post("/rules")
def create_rule(
    object_id: int = Form(...),
    utility_type: str = Form(...),
    mode: str = Form(...),
    base_area_mode: str = Form(...),
    tenant_id: str = Form(""),
    value_type: str = Form(""),
    value: str = Form(""),
    db: Session = Depends(get_db),
):
    if db.get(PropertyObject, object_id) is None:
        raise HTTPException(status_code=404, detail="Объект не найден.")
    require_choice(utility_type, UTILITY_TYPES, "услуга")
    
    actual_tenant_id: Optional[int] = None
    if tenant_id and tenant_id.strip():
        try:
            actual_tenant_id = int(tenant_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Недопустимый ID арендатора.")

    if actual_tenant_id and db.get(Tenant, actual_tenant_id) is None:
        raise HTTPException(status_code=404, detail="Арендатор не найден.")

    rule = AllocationRule(
        object_id=object_id,
        utility_type=utility_type,
        mode=mode,
        base_area_mode=base_area_mode,
        tenant_id=actual_tenant_id,
        value_type=value_type or None,
        value=to_decimal(value),
    )
    db.add(rule)
    db.commit()
    # Recalculate draft periods
    draft_periods = db.scalars(select(BillingPeriod).where(BillingPeriod.status == "draft")).all()
    for p in draft_periods:
        try:
            recalculate_period(db, p.id)
        except Exception:
            pass
    return RedirectResponse("/charges", status_code=303)


@app.post("/charges/{charge_id}/edit")
def edit_charge(
    charge_id: int,
    object_id: int = Form(...),
    billing_period_id: int = Form(...),
    utility_type: str = Form(...),
    input_mode: str = Form(...),
    allocation_mode: str = Form(...),
    amount: str = Form(""),
    meter_from: str = Form(""),
    meter_to: str = Form(""),
    tariff: str = Form(""),
    unit_name: str = Form(""),
    comment: str = Form(""),
    db: Session = Depends(get_db),
):
    charge = db.get(UtilityCharge, charge_id)
    if charge is None:
        raise HTTPException(status_code=404, detail="Счет не найден.")
    if db.get(PropertyObject, object_id) is None:
        raise HTTPException(status_code=404, detail="Объект не найден.")
    if db.get(BillingPeriod, billing_period_id) is None:
        raise HTTPException(status_code=404, detail="Период не найден.")
    require_choice(utility_type, UTILITY_TYPES, "услуга")
    
    charge.object_id = object_id
    charge.billing_period_id = billing_period_id
    charge.utility_type = utility_type
    charge.input_mode = input_mode
    charge.allocation_mode = allocation_mode
    charge.amount = to_decimal(amount)
    charge.meter_from = to_decimal(meter_from)
    charge.meter_to = to_decimal(meter_to)
    charge.tariff = to_decimal(tariff)
    charge.unit_name = unit_name.strip() or None
    charge.comment = comment.strip() or None
    
    get_charge_amount(charge)
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/charges", status_code=303)


@app.post("/charges/{charge_id}/delete")
def delete_charge(charge_id: int, db: Session = Depends(get_db)):
    charge = db.get(UtilityCharge, charge_id)
    if charge is None:
        raise HTTPException(status_code=404, detail="Счет не найден.")
    utility_name = UTILITY_LABELS.get(charge.utility_type, charge.utility_type)
    move_to_trash(db, charge, "utility_charge", f"Счет: {utility_name} на объекте {charge.object.name} ({charge.amount or 0} руб.)")
    db.execute(delete(ChargeAllocation).where(ChargeAllocation.utility_charge_id == charge_id))
    db.delete(charge)
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/charges", status_code=303)


@app.post("/rules/{rule_id}/edit")
def edit_rule(
    rule_id: int,
    object_id: int = Form(...),
    utility_type: str = Form(...),
    mode: str = Form(...),
    base_area_mode: str = Form(...),
    tenant_id: str = Form(""),
    value_type: str = Form(""),
    value: str = Form(""),
    db: Session = Depends(get_db),
):
    rule = db.get(AllocationRule, rule_id)
    if rule is None:
        raise HTTPException(status_code=404, detail="Правило не найдено.")
    if db.get(PropertyObject, object_id) is None:
        raise HTTPException(status_code=404, detail="Объект не найден.")
    require_choice(utility_type, UTILITY_TYPES, "услуга")
    
    actual_tenant_id: Optional[int] = None
    if tenant_id and tenant_id.strip():
        try:
            actual_tenant_id = int(tenant_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Недопустимый ID арендатора.")

    if actual_tenant_id and db.get(Tenant, actual_tenant_id) is None:
        raise HTTPException(status_code=404, detail="Арендатор не найден.")
        
    rule.object_id = object_id
    rule.utility_type = utility_type
    rule.mode = mode
    rule.base_area_mode = base_area_mode
    rule.tenant_id = actual_tenant_id
    rule.value_type = value_type or None
    rule.value = to_decimal(value)
    
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/charges", status_code=303)


@app.post("/rules/{rule_id}/delete")
def delete_rule(rule_id: int, db: Session = Depends(get_db)):
    rule = db.get(AllocationRule, rule_id)
    if rule is None:
        raise HTTPException(status_code=404, detail="Правило не найдено.")
    utility_name = UTILITY_LABELS.get(rule.utility_type, rule.utility_type)
    tenant_name = rule.tenant.display_name if rule.tenant else "всех"
    move_to_trash(db, rule, "allocation_rule", f"Правило: {utility_name} на {rule.object.name} для {tenant_name}")
    db.delete(rule)
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/charges", status_code=303)


@app.post("/tariffs")
def create_tariff(
    object_id: int = Form(...),
    tenant_id: str = Form(""),
    utility_type: str = Form(...),
    value: str = Form(...),
    unit_name: str = Form(""),
    start_date: date = Form(...),
    end_date: Optional[date] = Form(None),
    is_active: bool = Form(False),
    db: Session = Depends(get_db),
):
    if db.get(PropertyObject, object_id) is None:
        raise HTTPException(status_code=404, detail="Объект не найден.")
        
    actual_tenant_id: Optional[int] = None
    if tenant_id and tenant_id.strip():
        try:
            actual_tenant_id = int(tenant_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Недопустимый ID арендатора.")
        if db.get(Tenant, actual_tenant_id) is None:
            raise HTTPException(status_code=404, detail="Арендатор не найден.")

    allowed_types = ("rent",) + UTILITY_TYPES
    require_choice(utility_type, allowed_types, "тип тарифа")

    val = to_decimal(value, Decimal("0")) or Decimal("0")
    require_positive(val, "значение тарифа")

    tariff = Tariff(
        object_id=object_id,
        tenant_id=actual_tenant_id,
        utility_type=utility_type,
        value=val,
        unit_name=unit_name.strip() or None,
        start_date=start_date,
        end_date=end_date,
        is_active=is_active,
    )
    db.add(tariff)
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/charges", status_code=303)


@app.post("/tariffs/{tariff_id}/edit")
def edit_tariff(
    tariff_id: int,
    object_id: int = Form(...),
    tenant_id: str = Form(""),
    utility_type: str = Form(...),
    value: str = Form(...),
    unit_name: str = Form(""),
    start_date: date = Form(...),
    end_date: Optional[date] = Form(None),
    is_active: bool = Form(False),
    db: Session = Depends(get_db),
):
    tariff = db.get(Tariff, tariff_id)
    if tariff is None:
        raise HTTPException(status_code=404, detail="Тариф не найден.")
    if db.get(PropertyObject, object_id) is None:
        raise HTTPException(status_code=404, detail="Объект не найден.")
        
    actual_tenant_id: Optional[int] = None
    if tenant_id and tenant_id.strip():
        try:
            actual_tenant_id = int(tenant_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Недопустимый ID арендатора.")
        if db.get(Tenant, actual_tenant_id) is None:
            raise HTTPException(status_code=404, detail="Арендатор не найден.")

    allowed_types = ("rent",) + UTILITY_TYPES
    require_choice(utility_type, allowed_types, "тип тарифа")

    val = to_decimal(value, Decimal("0")) or Decimal("0")
    require_positive(val, "значение тарифа")

    tariff.object_id = object_id
    tariff.tenant_id = actual_tenant_id
    tariff.utility_type = utility_type
    tariff.value = val
    tariff.unit_name = unit_name.strip() or None
    tariff.start_date = start_date
    tariff.end_date = end_date
    tariff.is_active = is_active
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/charges", status_code=303)


@app.post("/tariffs/{tariff_id}/delete")
def delete_tariff(tariff_id: int, db: Session = Depends(get_db)):
    tariff = db.get(Tariff, tariff_id)
    if tariff is None:
        raise HTTPException(status_code=404, detail="Тариф не найден.")
    utility_name = "Аренда" if tariff.utility_type == "rent" else UTILITY_LABELS.get(tariff.utility_type, tariff.utility_type)
    tenant_name = tariff.tenant.display_name if tariff.tenant else "всех"
    move_to_trash(db, tariff, "tariff", f"Тариф: {utility_name} на {tariff.object.name} для {tenant_name} ({tariff.value})")
    db.delete(tariff)
    db.commit()
    recalculate_drafts(db)
    return RedirectResponse("/charges", status_code=303)


@app.get("/billing")
def billing_page(request: Request, db: Session = Depends(get_db)):
    periods = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).all()
    selected_id = request.query_params.get("period_id")
    selected_period = db.get(BillingPeriod, int(selected_id)) if selected_id else (periods[0] if periods else None)
    allocations = []
    totals: dict[str, Decimal] = {}
    if selected_period:
        allocations = db.scalars(
            select(ChargeAllocation)
            .options(
                joinedload(ChargeAllocation.object),
                joinedload(ChargeAllocation.tenant),
                joinedload(ChargeAllocation.utility_charge),
            )
            .where(ChargeAllocation.billing_period_id == selected_period.id)
            .order_by(ChargeAllocation.object_id, ChargeAllocation.tenant_id)
        ).all()
        for allocation in allocations:
            totals.setdefault(allocation.tenant.display_name, Decimal("0.00"))
            totals[allocation.tenant.display_name] += Decimal(allocation.amount)
    return render(
        request,
        "billing.html",
        db,
        periods=periods,
        selected_period=selected_period,
        allocations=allocations,
        totals={k: quantize_money(v) for k, v in totals.items()},
    )


@app.post("/billing/periods")
def create_billing_period(
    period_type: str = Form(...),
    month_label: str = Form(""),
    start_date: date = Form(...),
    end_date: date = Form(...),
    db: Session = Depends(get_db),
):
    if end_date < start_date:
        raise HTTPException(status_code=400, detail="Дата окончания не может быть раньше даты начала.")
    period = BillingPeriod(
        period_type=period_type,
        month_label=month_label.strip() or None,
        start_date=start_date,
        end_date=end_date,
    )
    db.add(period)
    db.commit()
    db.refresh(period)
    return RedirectResponse(f"/billing?period_id={period.id}", status_code=303)


@app.post("/billing/{period_id}/calculate")
def calculate_period(request: Request, period_id: int, db: Session = Depends(get_db)):
    warning = None
    try:
        recalculate_period(db, period_id)
    except ValueError as exc:
        # Расчёт завершён, но есть предупреждения (например, нет тарифа или нет арендаторов)
        # Отображаем их как warning, не блокируя результат
        warning = str(exc)
    return render_billing_page(request, db, selected_period_id=period_id, error=warning,
                                status_code=200 if warning else 200)


@app.post("/billing/allocations/{allocation_id}")
def update_allocation(
    allocation_id: int,
    amount: str = Form(...),
    db: Session = Depends(get_db),
):
    allocation = db.get(ChargeAllocation, allocation_id)
    if allocation is None:
        raise HTTPException(status_code=404, detail="Начисление не найдено.")
    period = db.get(BillingPeriod, allocation.billing_period_id)
    if period and period.status == "closed":
        raise HTTPException(status_code=400, detail="Закрытый период нельзя редактировать.")
    allocation.amount = to_decimal(amount, Decimal("0")) or Decimal("0")
    allocation.manual_override = True
    db.commit()
    return RedirectResponse(f"/billing?period_id={allocation.billing_period_id}", status_code=303)


@app.post("/billing/{period_id}/close")
def close_billing_period(period_id: int, db: Session = Depends(get_db)):
    close_period(db, period_id)
    return RedirectResponse(f"/billing?period_id={period_id}", status_code=303)


@app.post("/billing/{period_id}/reopen")
def reopen_billing_period(period_id: int, db: Session = Depends(get_db)):
    period = db.get(BillingPeriod, period_id)
    if period is None:
        raise HTTPException(status_code=404, detail="Период не найден.")
    period.status = "draft"
    period.closed_at = None
    db.commit()
    return RedirectResponse(f"/billing?period_id={period_id}", status_code=303)


@app.post("/billing/{period_id}/reset")
def reset_billing_period(period_id: int, db: Session = Depends(get_db)):
    period = db.get(BillingPeriod, period_id)
    if period is None:
        raise HTTPException(status_code=404, detail="Период не найден.")
    if period.status == "closed":
        raise HTTPException(status_code=400, detail="Закрытый период нельзя сбросить.")
    
    db.execute(delete(ChargeAllocation).where(ChargeAllocation.billing_period_id == period_id))
    period.status = "draft"
    db.commit()
    
    try:
        recalculate_period(db, period_id)
    except ValueError:
        pass
        
    return RedirectResponse(f"/billing?period_id={period_id}", status_code=303)


@app.get("/documents")
def documents_page(request: Request, db: Session = Depends(get_db)):
    periods = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).all()
    tenants = db.scalars(select(Tenant).order_by(Tenant.display_name)).all()
    grouped_periods = get_grouped_documents(db)
    return render(request, "documents.html", db, periods=periods, tenants=tenants, grouped_periods=grouped_periods)


@app.get("/guide")
def guide_page(request: Request, db: Session = Depends(get_db)):
    return render(request, "guide.html", db)


@app.get("/trash")
def trash_page(request: Request, db: Session = Depends(get_db)):
    items = db.scalars(select(TrashBin).order_by(TrashBin.deleted_at.desc())).all()
    formatted_items = []
    ENTITY_TYPE_LABELS = {
        "property_object": "Объект недвижимости",
        "tenant": "Арендатор",
        "lease_placement": "Размещение арендатора",
        "utility_charge": "Коммунальный счет",
        "allocation_rule": "Правило распределения",
        "tariff": "Тариф",
    }
    for item in items:
        formatted_items.append({
            "id": item.id,
            "display_name": item.display_name,
            "entity_type_label": ENTITY_TYPE_LABELS.get(item.entity_type, item.entity_type),
            "deleted_at": item.deleted_at.strftime("%d.%m.%Y %H:%M")
        })
    return render(request, "trash.html", db, items=formatted_items)


@app.post("/trash/{trash_id}/restore")
def restore_item(trash_id: int, db: Session = Depends(get_db)):
    item = db.get(TrashBin, trash_id)
    if item is None:
        raise HTTPException(status_code=404, detail="Запись в корзине не найдена.")
    
    data = json.loads(item.data_json)
    
    try:
        if item.entity_type == "property_object":
            obj = PropertyObject(
                id=data.get("id"),
                name=data.get("name"),
                address=data.get("address"),
                total_area=Decimal(data.get("total_area")),
                note=data.get("note")
            )
            db.add(obj)
        elif item.entity_type == "tenant":
            tenant = Tenant(
                id=data.get("id"),
                tenant_type=data.get("tenant_type"),
                display_name=data.get("display_name"),
                phone=data.get("phone"),
                note=data.get("note")
            )
            db.add(tenant)
        elif item.entity_type == "lease_placement":
            obj_exists = db.get(PropertyObject, data.get("object_id"))
            tenant_exists = db.get(Tenant, data.get("tenant_id"))
            if not obj_exists or not tenant_exists:
                raise HTTPException(status_code=400, detail="Невозможно восстановить: связанный объект или арендатор был удален насовсем.")
            
            placement = LeasePlacement(
                id=data.get("id"),
                object_id=data.get("object_id"),
                tenant_id=data.get("tenant_id"),
                rental_address=data.get("rental_address"),
                occupied_area=Decimal(data.get("occupied_area")),
                start_date=date.fromisoformat(data.get("start_date")),
                end_date=date.fromisoformat(data.get("end_date")) if data.get("end_date") else None,
                is_active=data.get("is_active")
            )
            db.add(placement)
        elif item.entity_type == "utility_charge":
            obj_exists = db.get(PropertyObject, data.get("object_id"))
            period_exists = db.get(BillingPeriod, data.get("billing_period_id"))
            if not obj_exists or not period_exists:
                raise HTTPException(status_code=400, detail="Невозможно восстановить: связанный объект или период был удален.")
                
            charge = UtilityCharge(
                id=data.get("id"),
                object_id=data.get("object_id"),
                billing_period_id=data.get("billing_period_id"),
                utility_type=data.get("utility_type"),
                input_mode=data.get("input_mode"),
                allocation_mode=data.get("allocation_mode"),
                amount=Decimal(data.get("amount")) if data.get("amount") else None,
                meter_from=Decimal(data.get("meter_from")) if data.get("meter_from") else None,
                meter_to=Decimal(data.get("meter_to")) if data.get("meter_to") else None,
                tariff=Decimal(data.get("tariff")) if data.get("tariff") else None,
                unit_name=data.get("unit_name"),
                comment=data.get("comment")
            )
            db.add(charge)
        elif item.entity_type == "allocation_rule":
            obj_exists = db.get(PropertyObject, data.get("object_id"))
            if not obj_exists:
                raise HTTPException(status_code=400, detail="Невозможно восстановить: связанный объект был удален.")
            
            rule = AllocationRule(
                id=data.get("id"),
                object_id=data.get("object_id"),
                utility_type=data.get("utility_type"),
                mode=data.get("mode"),
                base_area_mode=data.get("base_area_mode"),
                tenant_id=data.get("tenant_id"),
                value_type=data.get("value_type"),
                value=Decimal(data.get("value")) if data.get("value") else None,
                is_active=data.get("is_active")
            )
            db.add(rule)
        elif item.entity_type == "tariff":
            obj_exists = db.get(PropertyObject, data.get("object_id"))
            if not obj_exists:
                raise HTTPException(status_code=400, detail="Невозможно восстановить: связанный объект был удален.")
            
            tariff = Tariff(
                id=data.get("id"),
                object_id=data.get("object_id"),
                tenant_id=data.get("tenant_id"),
                utility_type=data.get("utility_type"),
                value=Decimal(data.get("value")),
                unit_name=data.get("unit_name"),
                start_date=date.fromisoformat(data.get("start_date")),
                end_date=date.fromisoformat(data.get("end_date")) if data.get("end_date") else None,
                is_active=data.get("is_active")
            )
            db.add(tariff)
            
        db.delete(item)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Ошибка восстановления: {str(e)}")
        
    recalculate_drafts(db)
    return RedirectResponse("/trash", status_code=303)


@app.post("/trash/{trash_id}/delete")
def permanent_delete_item(trash_id: int, db: Session = Depends(get_db)):
    item = db.get(TrashBin, trash_id)
    if item is None:
        raise HTTPException(status_code=404, detail="Запись в корзине не найдена.")
    db.delete(item)
    db.commit()
    return RedirectResponse("/trash", status_code=303)


@app.post("/trash/clear")
def clear_trash(db: Session = Depends(get_db)):
    db.execute(delete(TrashBin))
    db.commit()
    return RedirectResponse("/trash", status_code=303)



@app.post("/documents/generate")
def generate_documents(
    request: Request,
    billing_period_id: int = Form(...),
    tenant_id: str = Form(""),
    scope: str = Form(...),
    db: Session = Depends(get_db),
):
    period = db.get(BillingPeriod, billing_period_id)
    if period is None:
        raise HTTPException(status_code=404, detail="Период не найден.")
        
    actual_tenant_id: Optional[int] = None
    if tenant_id and tenant_id.strip():
        try:
            actual_tenant_id = int(tenant_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Недопустимый ID арендатора.")

    try:
        if scope == "all":
            generate_register_xlsx(db, period)
            tenant_ids = db.scalars(
                select(ChargeAllocation.tenant_id).where(ChargeAllocation.billing_period_id == period.id).distinct()
            ).all()
            for current_tenant_id in tenant_ids:
                tenant = db.get(Tenant, current_tenant_id)
                if tenant:
                    generate_invoice_docx(db, period, tenant)
                    generate_act_docx(db, period, tenant)
                    generate_invoice_xlsx(db, period, tenant)
                    generate_act_xlsx(db, period, tenant)
        else:
            tenant = db.get(Tenant, actual_tenant_id) if actual_tenant_id else None
            if tenant is None:
                raise ValueError("Для режима 'Для одного арендатора' необходимо выбрать конкретного арендатора.")
            generate_invoice_docx(db, period, tenant)
            generate_act_docx(db, period, tenant)
            generate_invoice_xlsx(db, period, tenant)
            generate_act_xlsx(db, period, tenant)
            generate_register_xlsx(db, period, tenant)
    except ValueError as exc:
        return render_documents_page(
            request,
            db,
            error=str(exc),
            form_state={
                "billing_period_id": billing_period_id,
                "tenant_id": tenant_id,
                "scope": scope,
            },
            status_code=400,
        )
    return RedirectResponse("/documents", status_code=303)


@app.get("/api/periods/{period_id}/allocations")
def api_period_allocations(period_id: int, db: Session = Depends(get_db)):
    """Возвращает JSON со списком арендаторов и их начислениями за период."""
    period = db.get(BillingPeriod, period_id)
    if not period:
        raise HTTPException(status_code=404, detail="Период не найден")

    UTILITY_LABELS = {
        "heat": "Тепло",
        "electricity": "Электричество",
        "water": "Водоснабжение",
        "cleaning": "Уборка",
        "other": "Прочее",
    }

    allocs = db.scalars(
        select(ChargeAllocation)
        .options(
            joinedload(ChargeAllocation.tenant),
            joinedload(ChargeAllocation.utility_charge),
        )
        .where(ChargeAllocation.billing_period_id == period_id)
        .order_by(ChargeAllocation.tenant_id)
    ).all()

    by_tenant: dict = defaultdict(list)
    for a in allocs:
        by_tenant[a.tenant_id].append(a)

    result = []
    for tenant_id, items in by_tenant.items():
        tenant = items[0].tenant
        services: dict = defaultdict(float)
        for a in items:
            # utility_type живёт в связанном UtilityCharge, не в ChargeAllocation
            utype = a.utility_charge.utility_type if a.utility_charge else "other"
            label = UTILITY_LABELS.get(utype, utype)
            services[label] += float(a.amount or 0)
        result.append({
            "tenant_id": tenant_id,
            "tenant_name": tenant.display_name if tenant else f"ID {tenant_id}",
            "total": round(sum(services.values()), 2),
            "services": dict(services),
        })

    return JSONResponse(result)


@app.post("/documents/generate-selected")
def generate_documents_selected(
    request: Request,
    billing_period_id: int = Form(...),
    tenant_ids: List[str] = Form(default=[]),
    db: Session = Depends(get_db),
):
    """Генерирует документы для выбранных арендаторов из списка начислений."""
    period = db.get(BillingPeriod, billing_period_id)
    if period is None:
        raise HTTPException(status_code=404, detail="Период не найден.")

    if not tenant_ids:
        return RedirectResponse("/documents", status_code=303)

    try:
        for tid_str in tenant_ids:
            try:
                tenant = db.get(Tenant, int(tid_str))
            except ValueError:
                continue
            if tenant:
                generate_invoice_docx(db, period, tenant)
                generate_act_docx(db, period, tenant)
                generate_invoice_xlsx(db, period, tenant)
                generate_act_xlsx(db, period, tenant)
                generate_register_xlsx(db, period, tenant)

        generate_register_xlsx(db, period)
    except ValueError as exc:
        return render_documents_page(
            request,
            db,
            error=str(exc),
            form_state={
                "billing_period_id": billing_period_id,
            },
            status_code=400,
        )
    return RedirectResponse("/documents", status_code=303)


def _get_clean_document_filename(doc: GeneratedDocument) -> str:
    p = doc.billing_period
    period_str = p.month_label or f"{p.start_date.isoformat()}_{p.end_date.isoformat()}"
    tenant_str = ""
    if doc.tenant:
        tenant_str = f"_{_safe_name(doc.tenant.display_name)}"
    
    ext = "docx"
    if doc.document_type.endswith("_xlsx") or doc.document_type == "register":
        ext = "xlsx"
        
    if doc.document_type == "register":
        if doc.tenant_id:
            return f"Реестр{tenant_str}_{period_str}.xlsx"
        return f"Реестр_начислений_{period_str}.xlsx"
        
    doc_type_label = "Счет" if "invoice" in doc.document_type else "Акт"
    return f"{doc_type_label}{tenant_str}_{period_str}.{ext}"


@app.get("/documents/{document_id}/download")
def download_document(document_id: int, db: Session = Depends(get_db)):
    document = db.scalar(
        select(GeneratedDocument)
        .options(joinedload(GeneratedDocument.tenant), joinedload(GeneratedDocument.billing_period))
        .where(GeneratedDocument.id == document_id)
    )
    if document is None:
        raise HTTPException(status_code=404, detail="Документ не найден.")
    clean_name = _get_clean_document_filename(document)
    return FileResponse(document.file_path, filename=clean_name)


@app.get("/documents/periods/{period_id}/zip")
def download_period_zip(period_id: int, db: Session = Depends(get_db)):
    period = db.get(BillingPeriod, period_id)
    if period is None:
        raise HTTPException(status_code=404, detail="Период не найден.")
        
    # Get all documents ordered by created_at desc so the newest are processed first
    documents = db.scalars(
        select(GeneratedDocument)
        .options(joinedload(GeneratedDocument.tenant), joinedload(GeneratedDocument.billing_period))
        .where(GeneratedDocument.billing_period_id == period.id)
        .order_by(GeneratedDocument.created_at.desc())
    ).all()
    
    if not documents:
        raise HTTPException(status_code=400, detail="Нет документов для скачивания за этот период.")
        
    # Deduplicate: only keep the newest document for each combination of (document_type, tenant_id)
    latest_docs = {}
    for doc in documents:
        key = (doc.document_type, doc.tenant_id)
        if key not in latest_docs:
            latest_docs[key] = doc
            
    final_docs = list(latest_docs.values())
        
    temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
    temp_zip.close()
    
    with zipfile.ZipFile(temp_zip.name, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for doc in final_docs:
            file_path = Path(doc.file_path)
            if file_path.exists():
                clean_name = _get_clean_document_filename(doc)
                zip_file.write(file_path, arcname=clean_name)
                
    period_label = period.month_label or f"{period.start_date}_{period.end_date}"
    zip_name = f"documents_{period_label}.zip"
    
    return FileResponse(
        temp_zip.name,
        filename=zip_name,
        media_type="application/zip"
    )


@app.get("/import/template")
def download_import_template():
    from openpyxl import Workbook
    workbook = Workbook()
    
    # Sheet 1: Objects
    sheet_obj = workbook.active
    sheet_obj.title = "Объекты"
    sheet_obj.append(["Название", "Адрес", "Площадь, кв.м", "Примечание"])
    sheet_obj.append(["БЦ Север", "Москва, ул. Ленина, д. 5", 500.50, "Главный корпус"])
    sheet_obj.append(["Склад Юг", "Москва, ул. Южная, д. 12", 250.00, "Теплый склад"])
    
    # Sheet 2: Tenants
    sheet_tenant = workbook.create_sheet(title="Арендаторы")
    sheet_tenant.append(["Тип (ИП/ООО)", "ФИО / Наименование", "Телефон", "Примечание"])
    sheet_tenant.append(["ООО", "Ромашка", "+7 (999) 111-22-33", "Офис 101"])
    sheet_tenant.append(["ИП", "Иванов Иван Иванович", "+7 (999) 555-44-33", "Офис 102"])
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    temp_file.close()
    workbook.save(temp_file.name)
    
    return FileResponse(
        temp_file.name,
        filename="import_template.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.post("/import")
def import_excel_data(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    # 0. Create database backup in JSON before running import
    make_db_backup(db)
    
    from openpyxl import load_workbook
    try:
        contents = file.file.read()
        buffer = io.BytesIO(contents)
        workbook = load_workbook(buffer, data_only=True)
        
        objects_created = 0
        tenants_created = 0
        
        # 1. Parse objects
        if "Объекты" in workbook.sheetnames:
            sheet = workbook["Объекты"]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                name, address, area, note = row[0], row[1] or "", row[2] or 0, row[3]
                
                existing_count = db.scalar(select(func.count(PropertyObject.id))) or 0
                if existing_count >= MAX_OBJECTS:
                    continue
                    
                obj = PropertyObject(
                    name=str(name).strip(),
                    address=str(address).strip(),
                    total_area=Decimal(str(area)),
                    note=str(note).strip() if note else None
                )
                db.add(obj)
                objects_created += 1
                
        # 2. Parse tenants
        if "Арендаторы" in workbook.sheetnames:
            sheet = workbook["Арендаторы"]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0] or not row[1]:
                    continue
                t_type, display_name, phone, note = row[0], row[1], row[2], row[3]
                
                existing_count = db.scalar(select(func.count(Tenant.id))) or 0
                if existing_count >= MAX_TENANTS:
                    continue
                    
                tenant = Tenant(
                    tenant_type=str(t_type).strip(),
                    display_name=str(display_name).strip(),
                    phone=str(phone).strip() if phone else None,
                    note=str(note).strip() if note else None
                )
                db.add(tenant)
                tenants_created += 1
                
        db.commit()
        
        # Recalculate draft periods
        draft_periods = db.scalars(select(BillingPeriod).where(BillingPeriod.status == "draft")).all()
        for p in draft_periods:
            try:
                recalculate_period(db, p.id)
            except Exception:
                pass
                
        objects_count = db.scalar(select(func.count(PropertyObject.id))) or 0
        tenants_count = db.scalar(select(func.count(Tenant.id))) or 0
        placements_count = db.scalar(select(func.count(LeasePlacement.id))) or 0
        
        latest_period = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).first()
        total_amount = Decimal("0.00")
        if latest_period:
            total_amount = db.scalar(
                select(func.sum(ChargeAllocation.amount))
                .where(ChargeAllocation.billing_period_id == latest_period.id)
            ) or Decimal("0.00")
            
        warnings = get_validation_warnings(db)
        from .paths import DATA_DIR
        has_backup = (DATA_DIR / "import_backup.json").exists()
        success_msg = f"Импорт завершен успешно! Создано объектов: {objects_created}, арендаторов: {tenants_created}."
        return render(
            request,
            "dashboard.html",
            db,
            objects_count=objects_count,
            tenants_count=tenants_count,
            placements_count=placements_count,
            latest_period=latest_period,
            total_amount=quantize_money(total_amount),
            success_msg=success_msg,
            warnings=warnings,
            has_backup=has_backup,
        )
    except Exception as e:
        objects_count = db.scalar(select(func.count(PropertyObject.id))) or 0
        tenants_count = db.scalar(select(func.count(Tenant.id))) or 0
        placements_count = db.scalar(select(func.count(LeasePlacement.id))) or 0
        latest_period = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).first()
        total_amount = Decimal("0.00")
        if latest_period:
            total_amount = db.scalar(
                select(func.sum(ChargeAllocation.amount))
                .where(ChargeAllocation.billing_period_id == latest_period.id)
            ) or Decimal("0.00")
            
        warnings = get_validation_warnings(db)
        from .paths import DATA_DIR
        has_backup = (DATA_DIR / "import_backup.json").exists()
        return render(
            request,
            "dashboard.html",
            db,
            objects_count=objects_count,
            tenants_count=tenants_count,
            placements_count=placements_count,
            latest_period=latest_period,
            total_amount=quantize_money(total_amount),
            error=f"Ошибка при импорте: {str(e)}",
            warnings=warnings,
            has_backup=has_backup,
            status_code=400,
        )


@app.post("/import/rollback")
def rollback_import(request: Request, db: Session = Depends(get_db)):
    success = restore_db_backup(db)
    
    objects_count = db.scalar(select(func.count(PropertyObject.id))) or 0
    tenants_count = db.scalar(select(func.count(Tenant.id))) or 0
    placements_count = db.scalar(select(func.count(LeasePlacement.id))) or 0
    latest_period = db.scalars(select(BillingPeriod).order_by(BillingPeriod.start_date.desc())).first()
    total_amount = Decimal("0.00")
    if latest_period:
        total_amount = db.scalar(
            select(func.sum(ChargeAllocation.amount))
            .where(ChargeAllocation.billing_period_id == latest_period.id)
        ) or Decimal("0.00")
        
    warnings = get_validation_warnings(db)
    msg = "Импорт успешно отменен! База данных восстановлена." if success else "Резервная копия не найдена."
    return render(
        request,
        "dashboard.html",
        db,
        objects_count=objects_count,
        tenants_count=tenants_count,
        placements_count=placements_count,
        latest_period=latest_period,
        total_amount=quantize_money(total_amount),
        success_msg=msg if success else None,
        error=None if success else msg,
        warnings=warnings,
        has_backup=False,
    )



